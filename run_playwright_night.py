#!/usr/bin/env python3
"""
🌙 Playwright 夜间跑机 — 模拟真人以图搜款
加强反检测 + 多策略上传 + 自动续跑

用法:
    python run_playwright_night.py --limit 200    # 跑全部
    python run_playwright_night.py --limit 5      # 测试5条
    python run_playwright_night.py --resume       # 断点续跑
"""
import sys
import os
import re
import json
import csv
import time
import random
import hashlib
import asyncio
import logging
from pathlib import Path
from typing import List, Dict
from dataclasses import dataclass, field, asdict

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

# ---- 日志 ----
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(ROOT / "output" / "playwright_night.log", mode="a", encoding="utf-8"),
    ],
)
logger = logging.getLogger("pw_night")

IMAGE_DIR = ROOT / "output" / "images"
IMAGE_DIR.mkdir(parents=True, exist_ok=True)
CHECKPOINT = ROOT / "output" / "pw_checkpoint.json"
OUTPUT_EXCEL = ROOT / "output" / "1688_playwright_result.xlsx"


# ============================================================
# 数据结构
# ============================================================
@dataclass
class ProductInfo:
    row: int
    image_url: str = ""
    image_local: str = ""
    title_ru: str = ""
    title_cn: str = ""
    category: str = ""
    price_rub: str = ""


@dataclass
class SearchResult:
    row: int = 0
    ozon_title: str = ""
    ozon_title_cn: str = ""
    ozon_price: str = ""
    status: str = ""  # success / no_result / captcha / error
    products: List[dict] = field(default_factory=list)


# ============================================================
# Step 1: 准备数据
# ============================================================
def load_products(limit: int = 0) -> List[ProductInfo]:
    """加载 Excel 数据 + 翻译 + 下载图片"""
    import openpyxl
    import requests as req

    # 翻译
    translations = {}
    csv_path = ROOT / "output" / "titles_for_translation.csv"
    if csv_path.exists():
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            for r in csv.reader(f):
                if len(r) >= 2 and r[0].strip() and r[1].strip():
                    translations[r[0].strip()] = r[1].strip()

    def find_translation(text: str) -> str:
        if text in translations:
            return translations[text]
        for key, val in translations.items():
            if key.startswith(text) or text.startswith(key):
                return val
        return ""

    wb = openpyxl.load_workbook(
        ROOT / "data" / "Seerfar-Product20260614_200.xlsx", data_only=True
    )
    ws = wb.active

    products = []
    for r in range(2, ws.max_row + 1):
        title_ru = str(ws.cell(row=r, column=4).value or "").strip()
        if not title_ru:
            continue
        products.append(ProductInfo(
            row=r - 1,
            image_url=str(ws.cell(row=r, column=2).value or "").strip(),
            title_ru=title_ru,
            title_cn=find_translation(title_ru),
            category=str(ws.cell(row=r, column=8).value or "").strip(),
            price_rub=str(ws.cell(row=r, column=10).value or "").strip(),
        ))
        if limit and len(products) >= limit:
            break
    wb.close()

    # 下载图片
    logger.info(f"准备 {len(products)} 个商品, 下载图片中...")
    session = req.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        "Referer": "https://www.ozon.ru/",
    })

    downloaded = 0
    for p in products:
        if not p.image_url:
            continue
        h = hashlib.md5(p.image_url.encode()).hexdigest()[:12]
        local = IMAGE_DIR / f"{h}.jpg"
        if local.exists() and local.stat().st_size > 100:
            p.image_local = str(local)
            downloaded += 1
            continue
        try:
            resp = session.get(p.image_url, timeout=15)
            resp.raise_for_status()
            with open(local, "wb") as f:
                f.write(resp.content)
            p.image_local = str(local)
            downloaded += 1
        except Exception as e:
            logger.debug(f"下载失败 {p.row}: {e}")

    logger.info(f"图片: {downloaded}/{len(products)}")
    return products


# ============================================================
# Step 2: Playwright 以图搜款 (增强版)
# ============================================================
class StealthSearcher:
    """加强反检测的 1688 以图搜款器"""

    def __init__(self):
        self._playwright = None
        self._browser = None
        self._context = None
        self._page = None
        self._captcha_count = 0

    async def start(self):
        from playwright.async_api import async_playwright

        self._playwright = await async_playwright().start()

        # 使用持久化上下文 — 保存 cookie/缓存
        user_data_dir = ROOT / "output" / "chrome_profile"
        user_data_dir.mkdir(parents=True, exist_ok=True)

        self._context = await self._playwright.chromium.launch_persistent_context(
            user_data_dir=str(user_data_dir),
            headless=True,
            viewport={"width": 1366, "height": 768},
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
            locale="zh-CN",
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-features=IsolateOrigins,site-per-process",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-web-security",
                "--disable-features=VizDisplayCompositor",
            ],
            ignore_default_args=["--enable-automation"],
        )

        # 注入反检测脚本
        await self._context.add_init_script("""
            // 隐藏自动化标志
            Object.defineProperty(navigator, 'webdriver', {get: () => false});
            Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
            Object.defineProperty(navigator, 'languages', {get: () => ['zh-CN', 'zh', 'en']});

            // 覆盖 chrome 对象
            window.chrome = {
                runtime: {},
                loadTimes: function() {},
                csi: function() {},
                app: {}
            };

            // 覆盖权限
            const originalQuery = window.navigator.permissions.query;
            window.navigator.permissions.query = (parameters) => (
                parameters.name === 'notifications' ?
                Promise.resolve({state: Notification.permission}) :
                originalQuery(parameters)
            );

            // 覆盖 headless 检测
            Object.defineProperty(navigator, 'hardwareConcurrency', {get: () => 8});
            Object.defineProperty(navigator, 'deviceMemory', {get: () => 8});
        """)

        self._page = await self._context.new_page()
        logger.info("🔧 浏览器就绪 (persistent + stealth)")

    async def search_by_image(self, image_path: str) -> List[dict]:
        """以图搜款 — 多策略尝试"""
        if not image_path or not Path(image_path).exists():
            return []

        logger.info(f"  📷 上传: {Path(image_path).name}")

        # 策略 1: 直接搜页面 → 找上传入口
        result = await self._try_search_page_upload(image_path)
        if result:
            return result

        # 策略 2: 首页 → 拍立淘入口
        result = await self._try_homepage_upload(image_path)
        if result:
            return result

        # 策略 3: 直接 POST 到 1688 搜索页
        result = await self._try_direct_search(image_path)
        if result:
            return result

        return []

    # ---- 策略 1: 搜索页上传 ----
    async def _try_search_page_upload(self, image_path: str) -> List[dict]:
        """直接打开 1688 搜索页，找图片上传按钮"""
        from playwright.async_api import TimeoutError as PWTO

        try:
            # 先打开搜索页（可能需要登录，试试看）
            resp = await self._page.goto(
                "https://s.1688.com/selloffer/offer_search.htm",
                wait_until="domcontentloaded",
                timeout=15000,
            )
            if "login" in self._page.url.lower():
                logger.debug("  策略1: 需要登录，跳过")
                return []

            await self._human_delay(1, 2)

            # 搜索页上可能有图片搜索按钮
            camera_btns = await self._page.query_selector_all(
                '[class*="camera"], [class*="image"], [class*="photo"], '
                'img[alt*="图"], [class*="pic-search"], .alisearch-camera'
            )
            if camera_btns:
                for btn in camera_btns:
                    try:
                        await btn.click()
                        await self._human_delay(0.5, 1.5)
                        break
                    except Exception:
                        continue

            # 找 file input
            file_input = await self._page.query_selector('input[type="file"]')
            if file_input:
                await file_input.set_input_files(image_path)
                await self._human_delay(3, 5)
                return await self._parse_results()

        except PWTO:
            logger.debug("  策略1: 超时")
        except Exception as e:
            logger.debug(f"  策略1: {e}")

        return []

    # ---- 策略 2: 首页拍立淘 ----
    async def _try_homepage_upload(self, image_path: str) -> List[dict]:
        """打开 1688 首页 → 找拍立淘入口"""
        try:
            resp = await self._page.goto(
                "https://www.1688.com/",
                wait_until="domcontentloaded",
                timeout=15000,
            )

            html = await self._page.content()
            if "captcha" in html.lower() or "interception" in html.lower():
                self._captcha_count += 1
                logger.debug(f"  策略2: 触发验证码 (#{self._captcha_count})")
                return []

            if "login" in self._page.url.lower():
                logger.debug("  策略2: 需要登录")
                return []

            await self._human_delay(0.5, 1.5)

            # 找以图搜货入口
            img_search = await self._page.query_selector(
                '.search-img-btn, .alisearch-camera, '
                'img[src*="camera"], img[alt*="拍照"], img[alt*="图片"], '
                'i[class*="camera"], [class*="image-search"]'
            )
            if img_search:
                await img_search.click()
                await self._human_delay(0.5, 1.5)

            # 找上传按钮
            file_input = await self._page.query_selector('input[type="file"]')
            if not file_input:
                # 看看有没有上传区域
                file_input = await self._page.query_selector(
                    '[class*="upload"], [class*="file-input"], '
                    'input[accept*="image"]'
                )

            if file_input:
                await file_input.set_input_files(image_path)
                await self._human_delay(3, 6)
                return await self._parse_results()

        except Exception as e:
            logger.debug(f"  策略2: {e}")

        return []

    # ---- 策略 3: 直连搜索 URL ----
    async def _try_direct_search(self, image_path: str) -> List[dict]:
        """用 requests 尝试直连 1688 以图搜图的内部 API"""
        # 这个方法对浏览器页面解析仍然靠 Playwright
        # 主要是换一个入口 URL
        try:
            # 试试 imageSearch 参数
            await self._page.goto(
                "https://s.1688.com/selloffer/offer_search.htm?imageSearch=1",
                wait_until="domcontentloaded",
                timeout=15000,
            )
            await self._human_delay(1, 2)

            file_input = await self._page.query_selector('input[type="file"]')
            if file_input:
                await file_input.set_input_files(image_path)
                await self._human_delay(3, 5)
                return await self._parse_results()

        except Exception as e:
            logger.debug(f"  策略3: {e}")

        return []

    # ---- 解析搜索结果 ----
    async def _parse_results(self) -> List[dict]:
        """从当前页面提取商品列表"""
        products = []

        try:
            await self._human_delay(1, 2)

            # 检查是否被拦截
            current_url = self._page.url
            if "login" in current_url or "punish" in current_url:
                return []

            html = await self._page.content()

            # 方法1: data 属性
            offer_ids = re.findall(r'data-offerid="(\d+)"', html)
            titles = re.findall(r'data-offer-title="([^"]*)"', html)
            prices = re.findall(r'data-offer-price="([^"]*)"', html)

            if offer_ids:
                for i in range(min(len(offer_ids), 12)):
                    price_text = prices[i] if i < len(prices) else ""
                    lo, hi = _parse_price(price_text)
                    products.append({
                        "offer_id": offer_ids[i],
                        "title": titles[i] if i < len(titles) else "",
                        "price_text": price_text,
                        "price_low": lo,
                        "price_high": hi,
                        "detail_url": f"https://detail.1688.com/offer/{offer_ids[i]}.html",
                    })

            # 方法2: CSS选择器
            if not products:
                cards = await self._page.query_selector_all(
                    '.space-offer-card, .sm-offer-item, '
                    'div[class*="offer-item"], li[class*="offer"]'
                )
                if not cards:
                    cards = await self._page.query_selector_all(
                        'div[class*="card"][class*="offer"]'
                    )

                for card in cards[:12]:
                    try:
                        title_el = await card.query_selector(
                            '.title, h3 a, [class*="subject"], a[href*="offer"]'
                        )
                        title = (await title_el.inner_text()).strip() if title_el else ""

                        price_el = await card.query_selector(
                            '.price, [class*="price-num"], [class*="price"]'
                        )
                        price_text = (await price_el.inner_text()).strip() if price_el else ""

                        link_el = await card.query_selector('a[href*="offer"]')
                        href = await link_el.get_attribute("href") if link_el else ""

                        if title or price_text:
                            lo, hi = _parse_price(price_text)
                            products.append({
                                "title": title,
                                "price_text": price_text,
                                "price_low": lo,
                                "price_high": hi,
                                "detail_url": f"https:{href}" if href.startswith("//") else href,
                            })
                    except Exception:
                        continue

            # 方法3: 暴力正则
            if not products:
                price_matches = re.findall(r'[¥￥]\s*([\d,.]+)', html)
                for i, pm in enumerate(price_matches[:12]):
                    lo, hi = _parse_price(pm)
                    products.append({
                        "title": f"商品{i+1}",
                        "price_text": f"¥{pm}",
                        "price_low": lo,
                        "price_high": hi,
                    })

        except Exception as e:
            logger.debug(f"  解析失败: {e}")

        return products

    async def close(self):
        if self._context:
            await self._context.close()
        if self._playwright:
            await self._playwright.stop()

    @staticmethod
    async def _human_delay(lo: float, hi: float):
        await asyncio.sleep(random.uniform(lo, hi))


def _parse_price(text: str) -> tuple:
    if not text:
        return 0.0, 0.0
    text = str(text).replace("¥", "").replace("￥", "").replace(",", "").strip()
    for sep in ("-", "~", "—", " "):
        if sep in text:
            parts = text.split(sep)
            try:
                a, b = float(parts[0].strip()), float(parts[-1].strip())
                return min(a, b), max(a, b)
            except ValueError:
                continue
    try:
        p = float(text)
        return p, p
    except ValueError:
        return 0.0, 0.0


# ============================================================
# Step 3: 主循环
# ============================================================
async def run(limit: int = 200, resume: bool = False):
    products = load_products(limit=limit)
    logger.info(f"共 {len(products)} 个商品待处理")

    # 加载断点
    results: List[SearchResult] = []
    done_rows = set()
    if resume and CHECKPOINT.exists():
        with open(CHECKPOINT, "r") as f:
            old = json.load(f)
            done_rows = {r["row"] for r in old if r.get("status") == "success"}
            results = [SearchResult(**r) for r in old]
            logger.info(f"断点续跑: 跳过 {len(done_rows)} 个已完成")

    searcher = StealthSearcher()
    await searcher.start()

    try:
        for i, p in enumerate(products):
            if p.row in done_rows:
                continue

            logger.info(f"\n{'─'*50}")
            logger.info(f"[{i+1}/{len(products)}] 第{p.row}行 {p.title_ru[:50]}")
            logger.info(f"  CN: {p.title_cn[:40] if p.title_cn else '(无翻译)'}")

            result = SearchResult(
                row=p.row,
                ozon_title=p.title_ru,
                ozon_title_cn=p.title_cn,
                ozon_price=p.price_rub,
            )

            # 以图搜款
            matches = await searcher.search_by_image(p.image_local)
            if matches:
                result.status = "success"
                result.products = matches[:12]
                logger.info(f"  ✅ 匹配 {len(matches)} 个商品")
                for m in matches[:3]:
                    logger.info(f"     {m.get('price_text', '?')} | {m.get('title', '?')[:40]}")
            else:
                if searcher._captcha_count >= 3:
                    result.status = "captcha"
                    logger.warning(f"  ⚠️ 验证码累计 {searcher._captcha_count} 次，暂停")
                else:
                    result.status = "no_result"
                    logger.info(f"  ⚠️ 无匹配")

            results.append(result)

            # 保存检查点
            with open(CHECKPOINT, "w", encoding="utf-8") as f:
                json.dump([asdict(r) for r in results], f, ensure_ascii=False, indent=2)

            # 间隔（模拟真人操作节奏）
            delay = random.uniform(5.0, 12.0)
            logger.info(f"  ⏱ 等待 {delay:.1f}s...")
            await asyncio.sleep(delay)

            # 验证码过多则停止
            if searcher._captcha_count >= 10:
                logger.error("验证码过多，停止运行")
                break

    finally:
        await searcher.close()

    # 输出统计
    success = sum(1 for r in results if r.status == "success")
    no_res = sum(1 for r in results if r.status == "no_result")
    captcha = sum(1 for r in results if r.status == "captcha")
    logger.info(f"\n{'='*50}")
    logger.info(f"完成! 成功:{success} | 无结果:{no_res} | 验证码:{captcha}")
    logger.info(f"检查点: {CHECKPOINT}")
    return results


# ============================================================
# 入口
# ============================================================
async def main():
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--limit", type=int, default=200)
    p.add_argument("--resume", action="store_true")
    args = p.parse_args()

    results = await run(limit=args.limit, resume=args.resume)

    # 打印摘要
    for r in results:
        if r.status == "success":
            for prod in r.products[:3]:
                logger.info(f"  #{r.row} | {prod.get('price_text','?')} | {prod.get('title','?')[:40]}")


if __name__ == "__main__":
    asyncio.run(main())
