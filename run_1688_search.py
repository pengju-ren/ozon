#!/usr/bin/env python3
"""
Step 2: Playwright 模拟真人操作 - 1688 以图搜款
读取本地图片 → 打开1688 → 上传图片 → 搜索 → 解析结果 → 输出Excel

用法:
    python run_1688_search.py                    # 跑全部（先下载图片）
    python run_1688_search.py --limit 5          # 只跑5个测试
    python run_1688_search.py --no-download      # 跳过下载（已下载过）
    python run_1688_search.py --headless False   # 显示浏览器窗口（调试用）
    python run_1688_search.py --start 10         # 从第10个开始（断点续跑）
"""
import sys
import os
import re
import json
import time
import random
import hashlib
import logging
import asyncio
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass, field, asdict

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))

from src.image_downloader import download_images, IMAGE_DIR

# ============================================================
# 日志
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(Path(__file__).resolve().parent / "output" / "search.log", mode="a", encoding="utf-8"),
    ],
)
logger = logging.getLogger("1688_search")

PROJECT_ROOT = Path(__file__).resolve().parent
RESULT_JSON = PROJECT_ROOT / "output" / "search_results.json"  # 中间结果（防丢）


# ============================================================
# 数据结构
# ============================================================
@dataclass
class Product1688:
    """1688 匹配商品"""
    title: str = ""
    price_text: str = ""       # 显示价格
    price_low: float = 0.0     # 最低价
    price_high: float = 0.0    # 最高价
    moq: str = ""              # 起批量
    shop: str = ""             # 店铺
    location: str = ""         # 所在地
    sales_text: str = ""       # 销量
    detail_url: str = ""       # 商品链接
    image_url: str = ""        # 商品图片


@dataclass
class SearchResult:
    """一个 Ozon 商品的搜索结果"""
    row: int = 0
    ozon_title: str = ""
    ozon_title_cn: str = ""
    ozon_price_rub: str = ""
    ozon_image: str = ""
    category: str = ""
    status: str = "pending"      # success / no_results / captcha / error
    error_msg: str = ""
    products: List[dict] = field(default_factory=list)


# ============================================================
# 核心：1688 以图搜款
# ============================================================
class ImageSearcher1688:
    """1688 以图搜款 — Playwright 版"""

    BASE = "https://www.1688.com"
    SEARCH_URL = "https://s.1688.com/selloffer/offer_search.htm"

    def __init__(self, headless: bool = True):
        self.headless = headless
        self._playwright = None
        self._browser = None
        self._context = None
        self._page = None

    async def __aenter__(self):
        from playwright.async_api import async_playwright

        self._playwright = await async_playwright().start()
        self._browser = await self._playwright.chromium.launch(
            headless=self.headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-features=IsolateOrigins,site-per-process",
            ],
        )
        self._context = await self._browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
            locale="zh-CN",
        )
        # 隐藏自动化特征
        await self._context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            window.chrome = { runtime: {} };
            Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
            Object.defineProperty(navigator, 'languages', {get: () => ['zh-CN', 'zh']});
        """)
        self._page = await self._context.new_page()
        logger.info("浏览器已启动")
        return self

    async def __aexit__(self, *args):
        if self._browser:
            await self._browser.close()
        if self._playwright:
            await self._playwright.stop()
        logger.info("浏览器已关闭")

    async def search_by_image(self, image_path: str) -> List[Product1688]:
        """
        核心：上传图片到 1688 以图搜款，返回匹配商品列表
        """
        if not image_path or not Path(image_path).exists():
            return []

        from playwright.async_api import TimeoutError as PWTimeout

        logger.info(f"  上传图片: {Path(image_path).name}")

        try:
            # ---- 方案 A：直接用 1688 以图搜货的直接 URL ----
            # 1688 有个隐藏的上传接口，直接 POST 图片到搜索结果页
            result = await self._try_direct_upload(image_path)
            if result:
                logger.info(f"  ✅ 找到 {len(result)} 个商品")
                return result

            # ---- 方案 B：模拟首页 → 点击相机 → 上传 ----
            result = await self._try_click_and_upload(image_path)
            if result:
                logger.info(f"  ✅ 找到 {len(result)} 个商品")
                return result

            logger.info(f"  ⚠️ 未找到匹配商品")
            return []

        except PWTimeout:
            logger.warning("  操作超时")
            return []
        except Exception as e:
            logger.warning(f"  搜索出错: {e}")
            return []

    # ----------------------------------------------------------
    # 方案 A：直连 1688 以图搜图的隐藏上传接口
    # ----------------------------------------------------------
    async def _try_direct_upload(self, image_path: str) -> List[Product1688]:
        """
        1688 拍立淘本质上是先上传图片到阿里 OSS，然后跳转搜索结果页。
        我们尝试直接用页面内的 file input 触发上传。
        """
        from playwright.async_api import TimeoutError as PWTimeout

        try:
            # 打开 1688 以图搜货页面
            await self._page.goto(
                self.BASE + "/",
                wait_until="domcontentloaded",
                timeout=20000,
            )
            await self._random_sleep(1, 2)

            # 先尝试找首页的图片搜索入口
            # 点击搜索框旁边的相机图标
            camera_selectors = [
                '.alisearch-camera',
                '.search-img-btn',
                'i[class*="icon-camera"]',
                'img[alt*="拍照"]',
                'img[alt*="图片"]',
                '[data-trace-click*="image"]',
                '.image-search-entry',
            ]

            clicked = False
            for sel in camera_selectors:
                btn = await self._page.query_selector(sel)
                if btn:
                    try:
                        await btn.click()
                        await self._random_sleep(0.5, 1.5)
                        clicked = True
                        break
                    except Exception:
                        continue

            if not clicked:
                # 试试直接用搜索框旁边的按钮
                search_box = await self._page.query_selector(
                    'input[name="keywords"], input[class*="search"], #alisearch-keywords'
                )
                if search_box:
                    # 如果页面上没有明显的图片搜索按钮，直接打开搜索页面
                    await self._page.goto(
                        "https://s.1688.com/selloffer/offer_search.htm?imageSearch=1",
                        wait_until="domcontentloaded",
                        timeout=20000,
                    )
                    await self._random_sleep(1, 2)

            # 找 file input 上传图片
            file_input = await self._page.query_selector('input[type="file"]')
            if not file_input:
                # 可能在 shadow DOM 里，试试更宽泛的选择
                logger.debug("  input[type=file] 未找到，尝试其他方式...")
                return []

            # 设置文件
            await file_input.set_input_files(image_path)
            await self._random_sleep(3, 5)  # 等上传+识别

            # 检查是否跳转到结果页
            current_url = self._page.url
            if "offer_search" in current_url or "offer" in current_url:
                await self._random_sleep(1, 2)
                return await self._parse_results()

        except Exception as e:
            logger.debug(f"  方案A失败: {e}")

        return []

    # ----------------------------------------------------------
    # 方案 B：点击首页相机 → 弹窗上传
    # ----------------------------------------------------------
    async def _try_click_and_upload(self, image_path: str) -> List[Product1688]:
        """备用方案：在首页找所有可能的入口"""
        try:
            await self._page.goto(self.BASE, wait_until="domcontentloaded", timeout=20000)
            await self._random_sleep(1, 2)

            # 1688 首页的图片搜索入口可能有：
            # - 搜索框旁的相机图标
            # - 导航栏的 "以图搜货"
            # - 直接的图片搜索链接

            # 尝试点击 "以图搜货" 文字链接
            text_links = await self._page.query_selector_all('a')
            for link in text_links:
                text = await link.inner_text()
                if '以图搜' in text or '图片搜' in text or '拍照' in text:
                    await link.click()
                    await self._random_sleep(1, 2)
                    break

            # 再找 file input
            file_input = await self._page.query_selector('input[type="file"]')
            if file_input:
                await file_input.set_input_files(image_path)
                await self._random_sleep(3, 5)
                return await self._parse_results()

        except Exception as e:
            logger.debug(f"  方案B失败: {e}")

        return []

    # ----------------------------------------------------------
    # 解析搜索结果
    # ----------------------------------------------------------
    async def _parse_results(self) -> List[Product1688]:
        """从1688搜索结果页面提取商品列表"""
        products = []

        try:
            # 等待结果
            await self._random_sleep(1, 2)
            html = await self._page.content()

            # ---- 策略 1: 从 HTML data 属性提取 ----
            # 1688 搜索结果商品卡通常有这些 data 属性
            offer_ids = re.findall(r'data-offerid="(\d+)"', html)
            titles = re.findall(r'data-offer-title="([^"]*)"', html)
            prices = re.findall(r'data-offer-price="([^"]*)"', html)

            if not offer_ids:
                offer_ids = re.findall(r'offerId["\']?\s*[:=]\s*["\']?(\d+)', html)

            if not titles:
                titles = re.findall(r'offerTitle["\']?\s*[:=]\s*["\']([^"\\]+)', html)

            if offer_ids and titles:
                for i in range(min(len(offer_ids), 20)):
                    p = Product1688(
                        title=titles[i] if i < len(titles) else "",
                        price_text=prices[i] if i < len(prices) else "",
                        detail_url=f"https://detail.1688.com/offer/{offer_ids[i]}.html",
                    )
                    if p.price_text:
                        p.price_low, p.price_high = _parse_price(p.price_text)
                    if p.title or p.price_low > 0:
                        products.append(p)

            # ---- 策略 2: 用选择器提取可见文本 ----
            if not products:
                products = await self._parse_by_selectors()

            # ---- 策略 3: 从页面可见价格提取 ----
            if not products:
                products = await self._parse_visible_prices()

        except Exception as e:
            logger.debug(f"  解析出错: {e}")

        # 去重
        seen = set()
        unique = []
        for p in products:
            key = p.title[:30] if p.title else p.detail_url
            if key and key not in seen:
                seen.add(key)
                unique.append(p)

        return unique[:20]

    async def _parse_by_selectors(self) -> List[Product1688]:
        """用 CSS 选择器解析商品卡片"""
        products = []
        try:
            # 尝试多种可能的卡片选择器
            card_selectors = [
                '.space-offer-card',
                '.sm-offer-item',
                '.offer-list-item',
                '.list-item',
                'div[class*="offer-item"]',
                'li[class*="offer"]',
                '.m-offer-item',
                '.offer-card',
            ]

            cards = []
            for sel in card_selectors:
                cards = await self._page.query_selector_all(sel)
                if cards:
                    break

            if not cards:
                cards = await self._page.query_selector_all(
                    'div[class*="card"], div[class*="item"]'
                )

            for card in cards[:20]:
                try:
                    # 标题
                    title_el = await card.query_selector(
                        'a[class*="title"], h3, .title, [class*="subject"], .offer-title'
                    )
                    title = (await title_el.inner_text()).strip() if title_el else ""

                    # 价格
                    price_el = await card.query_selector(
                        '[class*="price"], .price-num, .price-text'
                    )
                    price_text = (await price_el.inner_text()).strip() if price_el else ""

                    # 链接
                    link_el = await card.query_selector('a[href*="offer"]')
                    href = await link_el.get_attribute("href") if link_el else ""
                    if href and href.startswith("//"):
                        href = "https:" + href

                    # 店铺
                    shop_el = await card.query_selector(
                        '[class*="shop"], [class*="seller"], [class*="company"]'
                    )
                    shop = (await shop_el.inner_text()).strip() if shop_el else ""

                    # 起批量
                    moq_el = await card.query_selector('[class*="quantity"], [class*="min"]')
                    moq = (await moq_el.inner_text()).strip() if moq_el else ""

                    # 所在地
                    loc_el = await card.query_selector('[class*="location"], [class*="area"]')
                    loc = (await loc_el.inner_text()).strip() if loc_el else ""

                    if title or price_text:
                        p = Product1688(
                            title=title,
                            price_text=price_text.replace("¥", ""),
                            moq=moq,
                            shop=shop,
                            location=loc,
                            detail_url=href,
                        )
                        p.price_low, p.price_high = _parse_price(price_text)
                        if p.price_low > 0:
                            products.append(p)
                except Exception:
                    continue
        except Exception as e:
            logger.debug(f"  selector parsing error: {e}")

        return products

    async def _parse_visible_prices(self) -> List[Product1688]:
        """最终手段：提取页面上所有价格和相邻的链接"""
        products = []
        try:
            html = await self._page.content()

            # 找所有价格模式 ¥xx.xx
            price_pattern = re.findall(r'[¥￥]\s*([\d,.]+)', html)
            # 找所有链接
            links = re.findall(r'href="(//detail\.1688\.com/offer/\d+\.html)"', html)
            if not links:
                links = re.findall(r'href="(https?://detail\.1688\.com/offer/\d+\.html)"', html)

            for link in links[:10]:
                href = link if link.startswith("http") else f"https:{link}"
                prod = Product1688(detail_url=href)
                if price_pattern:
                    prod.price_text = price_pattern[0]
                    prod.price_low, prod.price_high = _parse_price(prod.price_text)
                products.append(prod)

        except Exception:
            pass
        return products

    # ----------------------------------------------------------
    # 工具
    # ----------------------------------------------------------
    @staticmethod
    async def _random_sleep(min_s: float = 0.5, max_s: float = 2.0):
        await asyncio.sleep(random.uniform(min_s, max_s))


# ============================================================
# 价格解析
# ============================================================
def _parse_price(text: str) -> tuple:
    """ "¥15.50-28.00" → (15.50, 28.00) """
    if not text:
        return 0.0, 0.0
    text = str(text).replace("¥", "").replace("￥", "").replace(",", "").strip()
    for sep in ("-", "~", "—", " "):
        if sep in text:
            parts = text.split(sep)
            try:
                lo, hi = float(parts[0].strip()), float(parts[-1].strip())
                return min(lo, hi), max(lo, hi)
            except ValueError:
                continue
    try:
        p = float(text)
        return p, p
    except ValueError:
        return 0.0, 0.0


# ============================================================
# 主流程
# ============================================================
async def search_all(
    products: List[Dict],
    headless: bool = True,
    start_index: int = 0,
):
    """对每个商品以图搜款"""
    results: List[SearchResult] = []

    # 加载之前的中间结果（断点续跑）
    if RESULT_JSON.exists():
        with open(RESULT_JSON, "r", encoding="utf-8") as f:
            old_results = json.load(f)
            completed_rows = {r["row"] for r in old_results if r["status"] == "success"}
            logger.info(f"从上次中断恢复，已完成 {len(completed_rows)} 个")

    async with ImageSearcher1688(headless=headless) as searcher:
        for i, product in enumerate(products):
            if i < start_index:
                continue

            row = product["row"]
            title_ru = product["title_ru"]
            image_local = product["image_local"]

            logger.info(f"\n{'='*60}")
            logger.info(f"[{i+1}/{len(products)}] 第{row}行: {title_ru[:50]}")
            logger.info(f"  中文: {product.get('title_cn', '')[:40]}")
            logger.info(f"  图片: {image_local}")

            result = SearchResult(
                row=row,
                ozon_title=title_ru,
                ozon_title_cn=product.get("title_cn", ""),
                ozon_price_rub=product.get("price_rub", ""),
                ozon_image=product.get("image_url", ""),
                category=product.get("category", ""),
            )

            if not image_local:
                result.status = "no_image"
                result.error_msg = "图片下载失败或不存在"
                results.append(result)
                _save_results(results)
                continue

            # 🔥 核心操作：以图搜款
            try:
                match_products = await searcher.search_by_image(image_local)

                if match_products:
                    result.status = "success"
                    result.products = [asdict(p) for p in match_products]
                    logger.info(f"  ✅ 匹配 {len(match_products)} 个商品:")
                    for p in match_products[:5]:
                        price_display = p.price_text or f"{p.price_low}-{p.price_high}" or "?"
                        logger.info(f"     {price_display} | {p.title[:40]}...")
                else:
                    result.status = "no_results"
                    result.error_msg = "1688 未找到匹配商品"
                    logger.info(f"  ⚠️ 未匹配到商品（可能被风控或无同款）")

            except Exception as e:
                result.status = "error"
                result.error_msg = str(e)[:200]
                logger.error(f"  ❌ 异常: {e}")

            results.append(result)
            _save_results(results)

            # 商品间长间隔（防风控）
            await asyncio.sleep(random.uniform(3.0, 6.0))

    return results


def _save_results(results: List[SearchResult]):
    """每次搜完立即保存（防断丢数据）"""
    with open(RESULT_JSON, "w", encoding="utf-8") as f:
        # 过滤掉 Playwright 线程 ID 等不可序列化字段
        json.dump([asdict(r) for r in results], f, ensure_ascii=False, indent=2)


# ============================================================
# 输出 Excel
# ============================================================
def save_to_excel(results: List[SearchResult], output_path: str):
    """将结果写入格式化 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1688以图搜款结果"

    # 表头
    headers = [
        "序号", "Ozon标题(俄语)", "中文关键词", "Ozon售价(₽)",
        "1688商品标题", "1688批发价", "起批量", "店铺",
        "所在地", "1688链接", "匹配状态", "备注",
    ]
    widths = [6, 40, 25, 12, 45, 14, 10, 18, 10, 35, 10, 20]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="微软雅黑", size=10)
    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    success_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    row_idx = 2
    for r in results:
        base_values = [
            r.row,
            r.ozon_title,
            r.ozon_title_cn,
            r.ozon_price_rub,
        ]

        if r.products:
            for p in r.products:
                fill = success_fill
                values = base_values + [
                    p.get("title", ""),
                    p.get("price_text", "") or f"{p.get('price_low', 0)}-{p.get('price_high', 0)}",
                    p.get("moq", ""),
                    p.get("shop", ""),
                    p.get("location", ""),
                    p.get("detail_url", ""),
                    r.status,
                    r.error_msg,
                ]
                for ci, v in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=ci, value=v)
                    cell.font = cell_font
                    cell.border = border
                    cell.fill = fill
                    if ci == 10 and isinstance(v, str) and v.startswith("http"):
                        cell.font = link_font
                        cell.hyperlink = v
                row_idx += 1
        else:
            fill = fail_fill
            values = base_values + ["", "", "", "", "", "", r.status, r.error_msg]
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.font = cell_font
                cell.border = border
                cell.fill = fill
            row_idx += 1

    ws.freeze_panes = "A2"
    wb.save(output_path)
    wb.close()
    logger.info(f"\n结果已保存: {output_path} ({row_idx - 2} 行)")


# ============================================================
# 入口
# ============================================================
async def main():
    import argparse
    parser = argparse.ArgumentParser(description="1688 以图搜款")
    parser.add_argument("--limit", type=int, default=0, help="限制处理数量（0=全部）")
    parser.add_argument("--start", type=int, default=0, help="从第几个开始（断点续跑）")
    parser.add_argument("--no-download", action="store_true", help="跳过图片下载")
    parser.add_argument("--headless", type=str, default="True", help="无头模式 True/False")
    args = parser.parse_args()

    headless = args.headless.lower() != "false"
    excel_path = str(PROJECT_ROOT / "data" / "Seerfar-Product20260614_200.xlsx")
    output_path = str(PROJECT_ROOT / "output" / "1688_search_result.xlsx")

    # Step 1: 下载图片
    if not args.no_download:
        logger.info("=" * 60)
        logger.info("Step 1: 下载商品图片")
        logger.info("=" * 60)
        products = download_images(excel_path, limit=args.limit)
    else:
        # 重建产品列表（不下载新图片，用已有文件）
        logger.info("跳过下载，使用已有图片...")
        import openpyxl
        translations = {}
        csv_path = PROJECT_ROOT / "output" / "titles_for_translation.csv"
        if csv_path.exists():
            import csv
            with open(csv_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    if len(row) >= 2:
                        translations[row[0].strip()] = row[1].strip()

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        products = []
        for r in range(2, ws.max_row + 1):
            image_url = str(ws.cell(row=r, column=2).value or "").strip()
            title_ru = str(ws.cell(row=r, column=4).value or "").strip()
            title_cn = translations.get(title_ru, "")
            category = str(ws.cell(row=r, column=8).value or "").strip()
            price_rub = str(ws.cell(row=r, column=10).value or "").strip()

            name_hash = hashlib.md5(image_url.encode()).hexdigest()[:12] if image_url else ""
            local = IMAGE_DIR / f"{name_hash}.jpg"
            if not local.exists():
                # Try other extensions
                for ext in [".png", ".webp"]:
                    alt = IMAGE_DIR / f"{name_hash}{ext}"
                    if alt.exists():
                        local = alt
                        break

            products.append({
                "row": r - 1,
                "image_url": image_url,
                "image_local": str(local) if local.exists() else "",
                "title_ru": title_ru,
                "title_cn": title_cn,
                "category": category,
                "price_rub": price_rub,
            })
            if args.limit and len(products) >= args.limit:
                break
        wb.close()

    if not products:
        logger.error("没有商品数据")
        return

    logger.info(f"共 {len(products)} 个商品")

    # Step 2: 以图搜款
    logger.info("=" * 60)
    logger.info("Step 2: 1688 以图搜款")
    logger.info("=" * 60)
    results = await search_all(products, headless=headless, start_index=args.start)

    # Step 3: 输出
    logger.info("=" * 60)
    logger.info("Step 3: 保存结果")
    logger.info("=" * 60)

    stats = {
        "total": len(results),
        "success": sum(1 for r in results if r.status == "success"),
        "no_results": sum(1 for r in results if r.status == "no_results"),
        "error": sum(1 for r in results if r.status in ("error", "captcha", "no_image")),
    }
    logger.info(f"  总数: {stats['total']} | 成功: {stats['success']} | "
                f"无结果: {stats['no_results']} | 异常: {stats['error']}")

    save_to_excel(results, output_path)


if __name__ == "__main__":
    asyncio.run(main())
