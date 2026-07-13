#!/usr/bin/env python3
"""
Ozon → 1688 货源匹配系统 — 统一入口

3 种搜索途径:
  途径 1: onebound   — OneBound API（推荐，日500次免费）
  途径 2: playwright — 浏览器自动化模拟真人（免费）
  途径 3: official   — 1688 官方 API（需营业执照）

用法:
  # OneBound API（默认）
  python run.py --limit 3

  # Playwright 模拟（显示浏览器窗口调试）
  python run.py --backend playwright --limit 3 --headless False

  # 指定输入/输出文件
  python run.py -i data/my_products.xlsx -o output/my_result.xlsx

  # 断点续跑
  python run.py --start 50

  # 只用关键词搜索
  python run.py --mode keyword

  # 只用图片搜索
  python run.py --mode image

  # 仅翻译模式（检查翻译质量）
  python run.py --translate-only
"""
import sys
import json
import asyncio
import logging
import argparse
from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import List, Dict

import openpyxl

# 项目根
ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

# 日志
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(ROOT / "output" / "pipeline.log", mode="a", encoding="utf-8"),
    ],
)
logger = logging.getLogger("ozon1688")

# ============================================================
# 数据结构
# ============================================================
@dataclass
class SearchResult:
    row: int = 0
    ozon_title: str = ""
    ozon_title_cn: str = ""
    ozon_price_rub: str = ""
    ozon_image: str = ""
    category: str = ""
    brand: str = ""
    sales: str = ""
    weight: str = ""
    status: str = "pending"
    error: str = ""
    products: List[dict] = field(default_factory=list)


# ============================================================
# 数据加载
# ============================================================
def load_products(excel_path: str, limit: int = 0) -> List[dict]:
    """从 Excel 加载商品数据 + 翻译（含模糊匹配）"""
    import csv

    # 加载翻译
    translations = {}
    csv_path = ROOT / "output" / "titles_for_translation.csv"
    if csv_path.exists():
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if len(row) >= 2 and row[0].strip() and row[1].strip():
                    translations[row[0].strip()] = row[1].strip()
    logger.info(f"翻译词库: {len(translations)} 条")

    def lookup_translation(text: str) -> str:
        """查找翻译，先精确匹配，再前缀模糊匹配"""
        if not text:
            return ""
        # 1. 精确匹配
        if text in translations:
            return translations[text]
        # 2. 前缀匹配（Excel 文本可能被截断）
        for key, val in translations.items():
            if key.startswith(text) or text.startswith(key):
                return val
        return ""

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    products = []

    for r in range(2, ws.max_row + 1):
        title_ru = str(ws.cell(row=r, column=4).value or "").strip()
        if not title_ru:
            continue

        products.append({
            "row": r - 1,
            "image_url": str(ws.cell(row=r, column=2).value or "").strip(),
            "title_ru": title_ru,
            "title_cn": lookup_translation(title_ru),
            "detail_url": str(ws.cell(row=r, column=5).value or "").strip(),
            "sku": str(ws.cell(row=r, column=6).value or "").strip(),
            "brand": str(ws.cell(row=r, column=7).value or "").strip(),
            "category": str(ws.cell(row=r, column=8).value or "").strip(),
            "price_rub": str(ws.cell(row=r, column=10).value or "").strip(),
            "sales": str(ws.cell(row=r, column=11).value or "").strip(),
            "weight": str(ws.cell(row=r, column=28).value or "").strip(),
        })

        if limit and len(products) >= limit:
            break

    wb.close()
    logger.info(f"读取商品: {len(products)} 条")
    return products


# ============================================================
# 主流程
# ============================================================
async def run_pipeline(
    products: List[dict],
    backend: str = "onebound",
    mode: str = "both",
    headless: bool = True,
    start_index: int = 0,
):
    """执行搜索管道"""
    results: List[SearchResult] = []

    # 断点恢复
    result_json = ROOT / "output" / "search_results.json"
    completed = set()
    if result_json.exists():
        with open(result_json, "r", encoding="utf-8") as f:
            old = json.load(f)
            completed = {r["row"] for r in old if r.get("status") == "success"}
            logger.info(f"断点恢复: 已跳过 {len(completed)} 个已完成商品")

    # 初始化搜索器
    if backend == "onebound":
        from src.searcher_onebound import OneBoundSearcher
        searcher = OneBoundSearcher()
        logger.info("🛤️  途径: OneBound API")
    elif backend == "playwright":
        from src.searcher_playwright import PlaywrightSearcher
        searcher = PlaywrightSearcher(headless=headless)
        logger.info("🛤️  途径: Playwright 模拟真人")
    elif backend == "official":
        from src.searcher_1688_official import Alibaba1688Searcher
        searcher = Alibaba1688Searcher()
        logger.info("🛤️  途径: 1688 官方 API")
    elif backend == "search1688api":
        from src.searcher_search1688api import Search1688ApiSearcher
        searcher = Search1688ApiSearcher()
        logger.info("🛤️  途径: search1688api H5 API")
    else:
        logger.error(f"未知途径: {backend}。可选: onebound / playwright / official / search1688api")
        return []

    logger.info(f"🔍 模式: {mode}")
    logger.info(f"📦 总数: {len(products)} | 起始: {start_index}")

    async with searcher:
        for i, p in enumerate(products):
            if i < start_index:
                continue

            row = p["row"]
            if row in completed:
                logger.info(f"[{i+1}/{len(products)}] 第{row}行 ⏭️ 已跳过")
                continue

            title = p["title_ru"]
            cn = p["title_cn"]
            logger.info(f"\n{'='*60}")
            logger.info(f"[{i+1}/{len(products)}] 第{row}行")
            logger.info(f"  俄语: {title[:60]}")
            logger.info(f"  中文: {cn[:40] if cn else '(无翻译)'}")

            result = SearchResult(
                row=row,
                ozon_title=title,
                ozon_title_cn=cn,
                ozon_price_rub=p["price_rub"],
                ozon_image=p["image_url"],
                category=p["category"],
                brand=p["brand"],
                sales=p["sales"],
                weight=p["weight"],
            )

            all_matches = []

            # 图片搜索
            if mode in ("image", "both") and p["image_url"]:
                try:
                    matches = await searcher.search_by_image(p["image_url"])
                    for m in matches:
                        m.similarity = getattr(m, "similarity", 0.0)
                    all_matches.extend(matches)
                except Exception as e:
                    logger.warning(f"  图片搜索异常: {e}")

            # 关键词搜索
            if mode in ("keyword", "both") and cn:
                await asyncio.sleep(1)  # 间隔
                try:
                    matches = await searcher.search_by_keyword(cn)
                    all_matches.extend(matches)
                except Exception as e:
                    logger.warning(f"  关键词搜索异常: {e}")

            if all_matches:
                # 去重 + 按价格排序
                seen = set()
                unique = []
                for m in all_matches:
                    key = m.offer_id or m.title[:30]
                    if key and key not in seen:
                        seen.add(key)
                        unique.append(m)
                # 按最低价排序
                unique.sort(key=lambda x: x.price_low if x.price_low > 0 else 999999)
                result.products = [asdict(m) for m in unique[:10]]
                result.status = "success"
                logger.info(f"  ✅ 匹配 {len(unique)} 个商品:")
                for m in unique[:5]:
                    logger.info(f"     {m.display_price} | {m.title[:40]}")
            else:
                result.status = "no_results"
                logger.info(f"  ⚠️ 无匹配结果")

            results.append(result)
            _save_checkpoint(results)

            # 间隔
            await asyncio.sleep(1.5)

    return results


def _save_checkpoint(results: List[SearchResult]):
    """每搜完一个商品就保存"""
    with open(ROOT / "output" / "search_results.json", "w", encoding="utf-8") as f:
        json.dump([asdict(r) for r in results], f, ensure_ascii=False, indent=2)


# ============================================================
# 输出 Excel
# ============================================================
def save_excel(results: List[SearchResult], output_path: str):
    """格式化输出 Excel"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1688货源匹配"

    headers = [
        "序号", "Ozon标题(俄语)", "中文关键词", "Ozon售价₽",
        "Ozon销量", "品牌", "类目",
        "1688商品标题", "1688批发价", "起批量", "店铺",
        "1688链接", "相似度", "状态", "备注",
    ]
    widths = [6, 40, 22, 10, 8, 10, 16, 40, 14, 8, 14, 32, 8, 8, 16]

    # 样式
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="微软雅黑", size=10)
    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    row_idx = 2
    for r in results:
        base = [
            r.row, r.ozon_title, r.ozon_title_cn, r.ozon_price_rub,
            r.sales, r.brand, r.category,
        ]
        if r.products:
            for p in r.products:
                vals = base + [
                    p.get("title", ""),
                    p.get("price_text", "") or f"{p.get('price_low', 0)}-{p.get('price_high', 0)}",
                    p.get("min_order", p.get("moq", "")),
                    p.get("shop_name", p.get("shop", "")),
                    p.get("detail_url", ""),
                    f"{p.get('similarity', 0):.0%}" if p.get("similarity") else "-",
                    r.status,
                    r.error,
                ]
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=ci, value=v)
                    cell.font = cell_font
                    cell.border = border
                    cell.fill = green_fill
                    if ci == 12 and str(v).startswith("http"):
                        cell.font = link_font
                        cell.hyperlink = str(v)
                row_idx += 1
        else:
            vals = base + ["", "", "", "", "", "", r.status, r.error]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.font = cell_font
                cell.border = border
                cell.fill = red_fill
            row_idx += 1

    ws.freeze_panes = "A2"
    wb.save(output_path)
    wb.close()
    logger.info(f"📊 结果已保存: {output_path} ({row_idx - 2} 行)")


# ============================================================
# CLI
# ============================================================
async def main():
    parser = argparse.ArgumentParser(
        description="Ozon → 1688 货源匹配（3种途径可选）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python run.py --backend onebound --limit 3      # OneBound API 测试3条
  python run.py --backend playwright --limit 3    # Playwright 测试3条
  python run.py --backend official --limit 3      # 1688官方API（需营业执照）
  python run.py --backend search1688api --limit 3 # H5 API 测试3条
  python run.py --mode keyword --limit 10         # 仅关键词搜索
  python run.py --translate-only                  # 检查翻译质量
        """,
    )
    parser.add_argument("--backend", choices=["onebound", "playwright", "official", "search1688api"],
                        default="onebound", help="搜索途径 (default: onebound)")
    parser.add_argument("--mode", choices=["image", "keyword", "both"],
                        default="both", help="搜索模式 (default: both)")
    parser.add_argument("--limit", type=int, default=0, help="限制商品数(0=全部)")
    parser.add_argument("--start", type=int, default=0, help="起始索引(断点续跑)")
    parser.add_argument("--headless", type=str, default="True",
                        help="Playwright 无头模式 True/False")
    parser.add_argument("-i", "--input", default="", help="输入Excel路径")
    parser.add_argument("-o", "--output", default="", help="输出Excel路径")
    parser.add_argument("--translate-only", action="store_true", help="仅测试翻译")

    args = parser.parse_args()

    excel_path = args.input or str(ROOT / "data" / "Seerfar-Product20260614_200.xlsx")
    output_path = args.output or str(ROOT / "output" / "1688_match_result.xlsx")

    # --translate-only
    if args.translate_only:
        products = load_products(excel_path, limit=5)
        for p in products:
            logger.info(f"  [{p['row']}] {p['title_ru'][:60]}")
            logger.info(f"       → {p['title_cn'] or '(无翻译)'}")
        return

    # 主流程
    products = load_products(excel_path, limit=args.limit)
    if not products:
        logger.error("没有商品数据")
        return

    headless = args.headless.lower() != "false"

    results = await run_pipeline(
        products=products,
        backend=args.backend,
        mode=args.mode,
        headless=headless,
        start_index=args.start,
    )

    if results:
        save_excel(results, output_path)
        stats = {
            "total": len(results),
            "success": sum(1 for r in results if r.status == "success"),
            "no_results": sum(1 for r in results if r.status == "no_results"),
        }
        logger.info(f"\n{'='*60}")
        logger.info(f"完成! 成功:{stats['success']}/{stats['total']} | "
                    f"无结果:{stats['no_results']}")
        logger.info(f"输出: {output_path}")


if __name__ == "__main__":
    asyncio.run(main())
