#!/usr/bin/env python3
"""
单条测试脚本 — 用一张图片 + 标题 + 类目搜索1688，获取价格
用法:
    python test_single.py              # 测试第1条
    python test_single.py --row 5      # 测试第5条
"""
import sys
import json
import argparse
import requests
from pathlib import Path

# 加载 .env
from dotenv import load_dotenv
import os
load_dotenv(Path(__file__).resolve().parent / ".env")

API_KEY = os.getenv("ONEBOUND_API_KEY", "")
API_SECRET = os.getenv("ONEBOUND_API_SECRET", "")
BASE_URL = "https://api-gw.onebound.cn/1688"
CSV_FILE = Path(__file__).resolve().parent / "output" / "titles_for_translation.csv"

# 加载翻译 CSV
import csv
translations = {}
if CSV_FILE.exists():
    with open(CSV_FILE, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            if len(row) >= 2 and row[0].strip() and row[1].strip():
                translations[row[0].strip()] = row[1].strip()
    print(f"✅ 加载了 {len(translations)} 条翻译")

import openpyxl


def translate(text: str) -> str:
    """从CSV查找翻译，找不到返回空"""
    return translations.get(text.strip(), "")


def search_by_image(image_url: str, page: int = 1):
    """以图搜款"""
    url = f"{BASE_URL}/item_search_img"
    params = {
        "key": API_KEY,
        "secret": API_SECRET,
        "imgid": image_url,
        "page": page,
        "page_size": 20,
        "sort": "sale_desc",
    }
    print(f"\n📷 图片搜索 (page={page})...")
    return _call_api(url, params)


def search_by_keyword(keyword: str, page: int = 1):
    """关键词搜索"""
    url = f"{BASE_URL}/item_search"
    params = {
        "key": API_KEY,
        "secret": API_SECRET,
        "q": keyword,
        "page": page,
        "page_size": 20,
        "sort": "default",
    }
    print(f"\n🔍 关键词搜索 (page={page}): \"{keyword}\"")
    return _call_api(url, params)


def _call_api(url: str, params: dict):
    """调用API并打印原始响应"""
    try:
        resp = requests.get(url, params=params, timeout=30)
        print(f"   HTTP {resp.status_code}, 长度: {len(resp.content)} bytes")

        data = resp.json()

        # 打印完整响应（截断过长内容）
        pretty = json.dumps(data, indent=2, ensure_ascii=False)
        if len(pretty) > 3000:
            print(f"   响应前3000字符:\n{pretty[:3000]}...")
        else:
            print(f"   完整响应:\n{pretty}")

        # 尝试解析商品列表
        items = _extract_items(data)
        if items:
            print(f"\n--- 找到 {len(items)} 个商品 ---")
            for i, item in enumerate(items[:10]):
                price = item.get("price", "?")
                title = item.get("title", "?")[:60]
                shop = item.get("seller_nick", item.get("company", "?"))
                similarity = item.get("similarity", item.get("match_rate", ""))
                sim_str = f" 相似度:{similarity}" if similarity else ""
                print(f"  [{i+1}] ¥{price} | {title} | 店铺:{shop}{sim_str}")
        else:
            print("   ⚠️ 未解析到商品列表")

        return data

    except requests.exceptions.RequestException as e:
        print(f"   ❌ 请求失败: {e}")
        return None
    except ValueError as e:
        print(f"   ❌ JSON解析失败: {e}")
        return None


def _extract_items(data: dict) -> list:
    """尝试多种格式提取商品列表"""
    if not data:
        return []
    # 格式1: {"items": {"item": [...]}}
    if "items" in data and isinstance(data["items"], dict):
        return data["items"].get("item", [])
    # 格式2: {"result": {"items": [...]}}
    if "result" in data and isinstance(data["result"], dict):
        return data["result"].get("items", [])
    # 格式3: {"data": {"items": [...]}}
    if "data" in data and isinstance(data["data"], dict):
        return data["data"].get("items", [])
    # 格式4: list directly
    if isinstance(data, list):
        return data
    # 格式5: {"item": [...]}
    if "item" in data:
        return data["item"] if isinstance(data["item"], list) else [data["item"]]
    return []


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--row", type=int, default=1, help="测试第几行数据(1=第一行)")
    parser.add_argument("--no-image", action="store_true", help="跳过图片搜索")
    parser.add_argument("--no-keyword", action="store_true", help="跳过关键词搜索")
    parser.add_argument("--keyword", type=str, default="", help="手动指定搜索关键词")
    args = parser.parse_args()

    input_file = Path(__file__).resolve().parent / "data" / "Seerfar-Product20260614_200.xlsx"
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb.active

    excel_row = args.row + 1  # +1 for header
    if excel_row > ws.max_row:
        print(f"❌ 行号超出范围: 总共 {ws.max_row - 1} 条数据")
        wb.close()
        return

    # 读取 B(2)=图片, D(4)=标题, H(8)=类目
    image_url = ws.cell(row=excel_row, column=2).value or ""
    title_ru = ws.cell(row=excel_row, column=4).value or ""
    category = ws.cell(row=excel_row, column=8).value or ""
    price_rub = ws.cell(row=excel_row, column=10).value or ""
    sales = ws.cell(row=excel_row, column=11).value or ""

    wb.close()

    # 翻译
    title_cn = translate(title_ru)

    print("=" * 70)
    print(f"📦 商品 #{args.row}")
    print(f"   俄语标题: {title_ru}")
    print(f"   中文关键词: {title_cn or '(未匹配到翻译)'}")
    print(f"   类目: {category}")
    print(f"   图片URL: {image_url[:80]}...")
    print(f"   Ozon售价: {price_rub} | 销量: {sales}")
    print(f"   API Key: {API_KEY[:8]}*** | Secret: {'已配置' if API_SECRET else '未配置'}")
    print("=" * 70)

    # 确定搜索关键词
    keyword = args.keyword or title_cn
    if not keyword:
        # 从类目中提取中文部分
        if category and "\n" in category:
            keyword = category.split("\n")[0]
        else:
            keyword = title_ru  # fallback

    # 图片搜索
    if not args.no_image and image_url:
        search_by_image(image_url)

    # 关键词搜索
    if not args.no_keyword and keyword:
        search_by_keyword(keyword)

    print("\n✅ 测试完成")


if __name__ == "__main__":
    main()
