#!/usr/bin/env python3
"""
导出 Ozon 产品标题，用于手动翻译或批量预翻译
输出 CSV 格式: 俄语原文,中文翻译
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

INPUT_FILE = Path(__file__).resolve().parent.parent / "data" / "Seerfar-Product20260614_200.xlsx"
OUTPUT_FILE = Path(__file__).resolve().parent.parent / "output" / "titles_for_translation.csv"


def main():
    print(f"读取: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb.active

    titles = set()  # 去重
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=4).value  # 标题列
        if title:
            titles.add(str(title).strip())

    wb.close()
    print(f"共 {len(titles)} 个唯一标题")

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8-sig") as f:
        f.write("俄语原文,中文翻译\n")
        for title in sorted(titles):
            # 用引号包裹包含逗号的文本
            escaped = f'"{title}"' if "," in title else title
            f.write(f"{escaped},\n")

    print(f"已导出: {OUTPUT_FILE}")
    print(f"\n使用方式:")
    print(f"  1. 打开 {OUTPUT_FILE}")
    print(f"  2. 在'中文翻译'列填入对应的中文关键词")
    print(f"  3. 在 .env 文件中设置: CSV_TRANSLATION_FILE=output/titles_for_translation.csv")


if __name__ == "__main__":
    main()
