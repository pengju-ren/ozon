"""
Seerfar → 1688 利润核算流水线

读 CSV → 翻译标题 → H5 图搜 → 关键词过滤 → 利润核算 → Excel

用法:
  python profit_pipeline.py                     # 跑全部
  python profit_pipeline.py --limit 3           # 只跑前 3 个测试
  python profit_pipeline.py --csv output/downloads/Seerfar-Product_20260715.csv
"""
import sys, os, re, csv, json, time, asyncio, logging, argparse
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass, field, asdict

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from search1688api import Sync1688Session
from src.llm_filter import translate_and_filter, translate_title as llm_translate

# ------------------------------------------------------------------
# 配置
# ------------------------------------------------------------------
RUB_TO_CNY = 0.078
DOMESTIC_SHIPPING = 5.0
INTL_SHIPPING_PER_KG = 30.0
OZON_COMMISSION_RATE = 0.10
DEFAULT_WEIGHT_KG = 0.3
OUTPUT_DIR = ROOT / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("pipeline")


# ------------------------------------------------------------------
# 数据结构
# ------------------------------------------------------------------
@dataclass
class MatchResult:
    """一条 1688 匹配 + 利润核算"""
    offer_id: str = ""
    title: str = ""
    price_text: str = ""         # 1688 批发价原文
    price_cny: float = 0.0       # 成本价(元)
    shop_name: str = ""
    shop_url: str = ""
    detail_url: str = ""
    sales: int = 0
    relevance: int = 0           # 过滤分数
    llm_score: int = 0           # LLM 相关性评分 (0-10)
    llm_reason: str = ""         # LLM 判断理由

    # 利润
    revenue_cny: float = 0.0
    cost_1688: float = 0.0
    cost_domestic: float = 0.0
    cost_intl_shipping: float = 0.0
    cost_commission: float = 0.0
    total_cost: float = 0.0
    net_profit: float = 0.0
    margin_pct: float = 0.0


@dataclass
class ProductResult:
    """一个 Ozon 商品的完整结果"""
    row: int = 0
    title_ru: str = ""
    title_cn: str = ""
    brand: str = ""
    category: str = ""
    price_rub: str = ""
    margin_pct_ozon: str = ""
    image_url: str = ""
    matches: List[MatchResult] = field(default_factory=list)
    note: str = ""               # 备注（如"仅配件，无整机"）
    error: str = ""


# ------------------------------------------------------------------
# ------------------------------------------------------------------
# 利润计算
# ------------------------------------------------------------------
def calc_profit(ozon_price_rub: str, cost_price_cny: float,
                weight_kg: float = DEFAULT_WEIGHT_KG) -> dict:
    clean = str(ozon_price_rub).replace("₽", "").replace(" ", "").replace(",", ".").strip()
    try:
        price_rub = float(clean)
    except (ValueError, TypeError):
        price_rub = 0.0

    revenue = price_rub * RUB_TO_CNY
    intl_shipping = weight_kg * INTL_SHIPPING_PER_KG
    commission = revenue * OZON_COMMISSION_RATE
    total_cost = cost_price_cny + DOMESTIC_SHIPPING + intl_shipping + commission
    net = revenue - total_cost
    margin = (net / revenue * 100) if revenue > 0 else 0.0

    return {
        "revenue_cny": round(revenue, 2),
        "cost_1688": round(cost_price_cny, 2),
        "cost_domestic": DOMESTIC_SHIPPING,
        "cost_intl_shipping": round(intl_shipping, 2),
        "cost_commission": round(commission, 2),
        "total_cost": round(total_cost, 2),
        "net_profit": round(net, 2),
        "margin_pct": round(margin, 1),
    }


# ------------------------------------------------------------------
# CSV 读取
# ------------------------------------------------------------------
def read_csv(csv_path: Path) -> List[Dict]:
    rows = []
    with open(csv_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            rows.append(row)
    logger.info(f"读取 CSV: {len(rows)} 行 ({csv_path.name})")
    return rows


# ------------------------------------------------------------------
# 主流程
# ------------------------------------------------------------------
async def run_pipeline(csv_path: Path, limit: int = 0,
                        start: int = 0, min_relevance: int = 1):
    rows = read_csv(csv_path)

    # 按毛利率降序重排（CSV 可能不是正确排序）
    def _margin_val(r):
        m = str(r.get("毛利率", "0")).replace("%", "").strip()
        try: return float(m)
        except: return 0.0
    rows.sort(key=_margin_val, reverse=True)

    if limit > 0:
        rows = rows[start:start + limit]

    results: List[ProductResult] = []
    searcher = Sync1688Session(debug=False)
    searcher.__enter__()

    try:
        for idx, row in enumerate(rows):
            row_num = idx + start + 1
            title_ru = (row.get("标题") or "").strip()
            brand = (row.get("品牌") or "").strip()
            category = (row.get("类目") or "").strip()
            price_rub = (row.get("售价") or "").strip()
            margin_ozon = (row.get("毛利率") or "").strip()
            image_url = (row.get("主图") or "").strip()

            logger.info(f"\n{'='*50}")
            logger.info(f"[{row_num}/{len(rows)+start}] {title_ru[:60]}")
            logger.info(f"  品牌: {brand or '-'} | 类目: {category[:30] if category else '-'}")

            pr = ProductResult(
                row=row_num, title_ru=title_ru, brand=brand,
                category=category, price_rub=price_rub,
                margin_pct_ozon=margin_ozon, image_url=image_url,
            )

            # 1. 下载图片 + 缩放
            local_img = _get_local_image(image_url, title_ru, row_num)
            if not local_img:
                pr.error = "无图片"
                results.append(pr)
                continue

            # 2. H5 图搜
            try:
                raw_items = searcher.search_by_image(str(local_img))
                logger.info(f"  图搜: {len(raw_items)} 条")
            except Exception as e:
                logger.warning(f"  图搜失败: {e}")
                pr.error = f"图搜失败: {e}"
                results.append(pr)
                continue

            if not raw_items:
                pr.error = "图搜无结果"
                results.append(pr)
                continue

            # 3. LLM 翻译 + 过滤（一条龙）
            title_cn, passed = translate_and_filter(
                row, raw_items, top_n=20
            )
            pr.title_cn = title_cn
            logger.info(f"  翻译: {title_cn[:60]}")
            logger.info(f"  LLM过滤: {len(passed)} 通过 / {len(raw_items) - len(passed)} 淘汰")

            # 5. 计算利润
            for item in passed:
                d = item.get("data", {})
                price_info = d.get("priceInfo", {})
                price_str = str(price_info.get("price", "0")) if isinstance(price_info, dict) else "0"
                price_cny = _parse_first_price(price_str)

                shop_add = d.get("shopAddition", {})
                shop_name = shop_add.get("text", "") if isinstance(shop_add, dict) else ""
                shop_url = shop_add.get("shopLinkUrl", "") if isinstance(shop_add, dict) else ""
                offer_id = str(d.get("offerId", ""))

                profit = calc_profit(price_rub, price_cny)
                m = MatchResult(
                    offer_id=offer_id,
                    title=d.get("title", ""),
                    price_text=price_str,
                    price_cny=price_cny,
                    shop_name=shop_name,
                    shop_url=shop_url,
                    detail_url=f"https://detail.1688.com/offer/{offer_id}.html" if offer_id else "",
                    sales=int(d.get("saleQuantity", d.get("bookedCount", 0))),
                    relevance=item.get("_relevance_score", 0),
                    llm_score=item.get("_llm_score", 0),
                    llm_reason=item.get("_llm_reason", ""),
                    **profit,
                )
                pr.matches.append(m)

            # 判断是否仅有配件
            if pr.matches:
                max_llm = max(m.llm_score for m in pr.matches)
                if max_llm < 7:
                    pr.note = "仅配件，无整机同款"
                elif max_llm < 9:
                    pr.note = "配件为主"

            if passed:
                # 按净利润降序排
                pr.matches.sort(key=lambda x: x.net_profit, reverse=True)
            results.append(pr)

    finally:
        searcher.__exit__(None, None, None)

    # 保存
    _save_excel(results, csv_path.stem)
    _save_json(results, csv_path.stem)
    _print_summary(results)

    return results


# ------------------------------------------------------------------
# 辅助
# ------------------------------------------------------------------
def _get_local_image(image_url: str, title: str, row_num: int = 0) -> Optional[Path]:
    """获取本地图片：先检查已下载的，没有再下载"""
    import hashlib
    img_dir = ROOT / "output" / "images"
    img_dir.mkdir(parents=True, exist_ok=True)

    # 1. 按 URL hash 查
    if image_url.startswith("http"):
        h = hashlib.md5(image_url.encode()).hexdigest()[:12]
        for ext in (".jpg", ".jpeg", ".png", "_s.jpg"):
            p = img_dir / f"{h}{ext}"
            if p.exists() and p.stat().st_size > 100:
                return p

    # 2. 下载（使用 H5 兼容尺寸）
    if image_url.startswith("http"):
        try:
            import urllib.request
            from PIL import Image
            h = hashlib.md5(image_url.encode()).hexdigest()[:12]
            raw_path = img_dir / f"{h}_raw.jpg"
            urllib.request.urlretrieve(image_url, raw_path)
            # 缩放到 200px（兼容 1688 H5 API 限制）
            img = Image.open(raw_path)
            max_dim = max(img.size)
            if max_dim > 200:
                ratio = 200 / max_dim
                img = img.resize(
                    (int(img.size[0] * ratio), int(img.size[1] * ratio)),
                    Image.LANCZOS
                )
            final_path = img_dir / f"{h}.jpg"
            img.save(final_path, quality=85)
            if raw_path != final_path:
                raw_path.unlink()  # 删原图
            logger.info(f"  下载+缩放: {final_path.name} ({img.size[0]}x{img.size[1]})")
            return final_path
        except Exception as e:
            logger.warning(f"  图片下载失败: {e}")
    return None


def _parse_first_price(price_str: str) -> float:
    """解析价格 "3.50-28.00" 或 "1.55万" → 取最低值"""
    if not price_str:
        return 0.0
    p = str(price_str).replace("¥", "").replace("￥", "").replace(",", "").strip()

    # 处理 "1.55万" 格式
    if "万" in p:
        m = re.match(r'(\d+\.?\d*)', p)
        if m:
            return float(m.group(1)) * 10000

    # 取第一个数字
    m = re.match(r'(\d+\.?\d*)', p)
    if m:
        return float(m.group(1))
    return 0.0


# ------------------------------------------------------------------
# 输出
# ------------------------------------------------------------------
def _save_excel(results: List[ProductResult], prefix: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl 未安装，跳过 Excel 输出")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "利润核算"

    headers = [
        "序号", "Ozon标题", "翻译标题", "品牌", "类目", "Ozon售价₽",
        "1688商品", "1688价格", "成本价¥", "店铺", "店铺链接",
        "1688链接", "1688销量", "LLM分", "LLM理由",
        "收入¥", "1688成本", "国内运费", "国际运费", "佣金",
        "总成本¥", "净利润¥", "毛利率%", "Ozon毛利率%", "备注",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for pr in results:
        items = pr.matches if pr.matches else [None]
        for m in items:
            row = ws.max_row + 1
            vals = [
                pr.row, pr.title_ru[:80], pr.title_cn[:60], pr.brand, pr.category[:40], pr.price_rub,
                m.title[:80] if m else (pr.error or "无匹配"),
                m.price_text if m else "-",
                m.price_cny if m else 0,
                m.shop_name if m else "-",
                m.shop_url if m else "-",
                m.detail_url if m else "-",
                m.sales if m else 0,
                m.llm_score if m else 0,
                m.llm_reason[:40] if m else "",
                m.revenue_cny if m else 0,
                m.cost_1688 if m else 0,
                m.cost_domestic if m else 0,
                m.cost_intl_shipping if m else 0,
                m.cost_commission if m else 0,
                m.total_cost if m else 0,
                m.net_profit if m else 0,
                m.margin_pct if m else 0,
                pr.margin_pct_ozon,
                pr.note if m else (pr.error or pr.note or ""),
            ]
            for c, v in enumerate(vals, 1):
                ws.cell(row=row, column=c, value=v)

    path = OUTPUT_DIR / f"{prefix}_profit.xlsx"
    wb.save(path)
    logger.info(f"Excel 已保存: {path}")




def _save_json(results: List[ProductResult], prefix: str):
    path = OUTPUT_DIR / f"{prefix}_profit.json"
    data = []
    for pr in results:
        data.append({
            "row": pr.row, "title_ru": pr.title_ru, "title_cn": pr.title_cn,
            "brand": pr.brand, "category": pr.category, "price_rub": pr.price_rub,
            "error": pr.error,
            "matches": [asdict(m) for m in pr.matches],
        })
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    logger.info(f"JSON 已保存: {path}")


def _print_summary(results: List[ProductResult]):
    total = len(results)
    matched = sum(1 for r in results if r.matches)
    profitable = sum(
        1 for r in results
        for m in r.matches if m.net_profit > 0
    )
    logger.info(f"\n{'='*50}")
    logger.info(f"📊 汇总: {total} 个商品, {matched} 个有匹配, {profitable} 个有利可图")


# ------------------------------------------------------------------
# CLI
# ------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Seerfar → 1688 利润核算流水线")
    parser.add_argument("--limit", type=int, default=3, help="处理前 N 个商品（0=全部）")
    parser.add_argument("--start", type=int, default=0, help="起始位置")
    parser.add_argument("--csv", help="CSV 文件路径")
    parser.add_argument("--min-relevance", type=int, default=2,
                        help="过滤最低相关分数（默认 2 = 至少命中一个类目词或品牌）")
    args = parser.parse_args()

    csv_path = Path(args.csv) if args.csv else sorted(
        (OUTPUT_DIR / "downloads").glob("Seerfar-Product_*.csv"),
        reverse=True
    )[0] if list((OUTPUT_DIR / "downloads").glob("Seerfar-Product_*.csv")) else None

    if not csv_path or not csv_path.exists():
        logger.error(f"找不到 CSV 文件: {csv_path}")
        sys.exit(1)

    asyncio.run(run_pipeline(
        csv_path=csv_path,
        limit=args.limit,
        start=args.start,
        min_relevance=args.min_relevance,
    ))
