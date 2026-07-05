"""
主流程模块 - 读取Excel → 翻译 → 搜索1688 → 匹配 → 输出Excel
"""
import os
import re
import time
import logging
from pathlib import Path
from typing import List, Optional, Iterator
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from .config import (
    DATA_DIR,
    OUTPUT_DIR,
    RUB_TO_CNY,
    SEARCH_PAGES,
    TOP_MATCHES_PER_PRODUCT,
    COL_RANK, COL_IMAGE, COL_TITLE, COL_URL, COL_SKU,
    COL_BRAND, COL_CATEGORY, COL_PRICE, COL_SALES,
    COL_MARGIN, COL_WEIGHT, COL_VOLUME,
    CSV_TRANSLATION_FILE,
)
from .translator import Translator
from .searcher_1688 import Searcher1688, Product1688
from .matcher import Matcher, ProductMatch, MatchResult

logger = logging.getLogger(__name__)


class Pipeline:
    """Ozon → 1688 货源匹配主流程"""

    def __init__(
        self,
        input_file: str,
        output_file: str = "",
        searcher: Optional[Searcher1688] = None,
        translator: Optional[Translator] = None,
        matcher: Optional[Matcher] = None,
        start_row: int = 0,
        max_products: int = 0,
    ):
        """
        Args:
            input_file: 输入Excel文件路径
            output_file: 输出Excel文件路径
            searcher: 1688搜索器（可选，自动创建）
            translator: 翻译器（可选，自动创建）
            matcher: 匹配器（可选，自动创建）
            start_row: 起始行（0-based，不含表头），用于断点续跑
            max_products: 最大处理数量，0=全部
        """
        self.input_file = Path(input_file)
        self.output_file = Path(output_file) if output_file else OUTPUT_DIR / "1688_match_result.xlsx"
        self.searcher = searcher or Searcher1688()
        self.translator = translator or Translator(csv_path=CSV_TRANSLATION_FILE)
        self.matcher = matcher or Matcher(top_n=TOP_MATCHES_PER_PRODUCT)
        self.start_row = start_row
        self.max_products = max_products

        # 统计
        self.stats = {
            "total": 0,
            "translated": 0,
            "image_searched": 0,
            "keyword_searched": 0,
            "matched": 0,
            "errors": 0,
        }

    # ================================================================
    # 主流程
    # ================================================================

    def run(self, use_image_search: bool = True, use_keyword_search: bool = True):
        """
        执行主流程

        Args:
            use_image_search: 是否使用以图搜款
            use_keyword_search: 是否使用关键词搜索
        """
        logger.info("=" * 60)
        logger.info("Ozon → 1688 货源匹配系统")
        logger.info(f"输入文件: {self.input_file}")
        logger.info(f"输出文件: {self.output_file}")
        logger.info(f"图片搜索: {'启用' if use_image_search else '禁用'}")
        logger.info(f"关键词搜索: {'启用' if use_keyword_search else '禁用'}")
        logger.info("=" * 60)

        # Step 1: 读取Excel
        products = self._read_excel()
        self.stats["total"] = len(products)
        logger.info(f"读取到 {len(products)} 个Ozon商品")

        if not products:
            logger.error("没有读取到商品数据")
            return

        # Step 2: 翻译标题
        logger.info("开始翻译俄语标题...")
        titles = [p.get("title", "") for p in products]
        translated = self.translator.translate_batch(titles)
        for p, cn in zip(products, translated):
            p["title_cn"] = cn
        self.stats["translated"] = len([t for t in translated if t])
        logger.info(f"翻译完成: {self.stats['translated']}/{len(products)} 成功")

        # Step 3: 对每个商品搜索1688
        all_matches: List[ProductMatch] = []

        for i, product in enumerate(tqdm(products, desc="搜索1688")):
            try:
                match = self._process_one_product(
                    product, i,
                    use_image_search=use_image_search,
                    use_keyword_search=use_keyword_search,
                )
                all_matches.append(match)
            except Exception as e:
                logger.error(f"处理商品 #{product.get('rank', i)} 失败: {e}")
                self.stats["errors"] += 1
                # 创建空的匹配结果
                empty_match = ProductMatch(
                    ozon_rank=product.get("rank", i + 1),
                    ozon_title=product.get("title", ""),
                    ozon_title_cn=product.get("title_cn", ""),
                    ozon_price_rub=product.get("price_rub", ""),
                    ozon_price_cny=product.get("price_cny", 0.0),
                    ozon_sku=str(product.get("sku", "")),
                    ozon_image=product.get("image", ""),
                    ozon_sales=product.get("sales", 0),
                    ozon_margin=product.get("margin", ""),
                    brand=product.get("brand", ""),
                    category=product.get("category", ""),
                )
                all_matches.append(empty_match)

        # Step 4: 写入输出Excel
        logger.info(f"写入结果到: {self.output_file}")
        self._write_excel(all_matches, products)

        # Step 5: 打印统计
        logger.info("=" * 60)
        logger.info("流程完成!")
        logger.info(f"  总商品数: {self.stats['total']}")
        logger.info(f"  翻译成功: {self.stats['translated']}")
        logger.info(f"  图片搜索: {self.stats['image_searched']}")
        logger.info(f"  关键词搜索: {self.stats['keyword_searched']}")
        logger.info(f"  成功匹配: {self.stats['matched']}")
        logger.info(f"  错误: {self.stats['errors']}")
        logger.info(f"  输出文件: {self.output_file}")
        logger.info("=" * 60)

    # ================================================================
    # 单个商品处理
    # ================================================================

    def _process_one_product(
        self,
        product: dict,
        index: int,
        use_image_search: bool = True,
        use_keyword_search: bool = True,
    ) -> ProductMatch:
        """处理单个Ozon商品"""
        match = ProductMatch(
            ozon_rank=product.get("rank", index + 1),
            ozon_title=product.get("title", ""),
            ozon_title_cn=product.get("title_cn", ""),
            ozon_price_rub=product.get("price_rub", ""),
            ozon_price_cny=product.get("price_cny", 0.0),
            ozon_sku=str(product.get("sku", "")),
            ozon_image=product.get("image", ""),
            ozon_sales=product.get("sales", 0),
            ozon_margin=product.get("margin", ""),
            brand=product.get("brand", ""),
            category=product.get("category", ""),
        )

        image_results: List[Product1688] = []
        keyword_results: List[Product1688] = []

        # 图片搜索（优先）
        image_url = product.get("image", "")
        if use_image_search and image_url:
            self.stats["image_searched"] += 1
            image_results = self.searcher.search_by_image_all_pages(image_url)
            time.sleep(0.5)

        # 关键词搜索
        if use_keyword_search:
            self.stats["keyword_searched"] += 1
            cn_keyword = product.get("title_cn", "")
            if cn_keyword:
                keyword_results = self.searcher.search_by_keyword_all_pages(cn_keyword)
                time.sleep(0.5)

        # 合并排序
        merged = self.matcher.merge_and_rank(image_results, keyword_results)
        match.matches = merged

        if merged:
            self.stats["matched"] += 1

        return match

    # ================================================================
    # Excel 读取
    # ================================================================

    def _read_excel(self) -> List[dict]:
        """读取Ozon产品Excel"""
        wb = openpyxl.load_workbook(self.input_file, data_only=True)
        ws = wb.active
        logger.info(f"读取工作表: {ws.title}, {ws.max_row}行 x {ws.max_column}列")

        # 读取表头，建立列名→列号映射
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers[str(val).strip()] = col

        logger.debug(f"识别到的列: {list(headers.keys())}")

        # 确定数据范围
        start_data_row = 2 + self.start_row  # 跳过表头 + 偏移
        end_data_row = ws.max_row
        if self.max_products > 0:
            end_data_row = min(end_data_row, start_data_row + self.max_products - 1)

        products = []
        for row_idx in range(start_data_row, end_data_row + 1):
            product = self._parse_row(ws, row_idx, headers)
            if product and product.get("title"):
                products.append(product)

        wb.close()
        return products

    def _parse_row(self, ws, row_idx: int, headers: dict) -> Optional[dict]:
        """解析Excel中的一行数据"""
        def get_val(col_name):
            col_idx = headers.get(col_name)
            if col_idx is None:
                return None
            return ws.cell(row=row_idx, column=col_idx).value

        title = get_val(COL_TITLE)
        if not title:
            return None

        # 解析售价（格式如 "5312₽"）
        price_rub_str = str(get_val(COL_PRICE) or "")
        price_rub = self._parse_rub_price(price_rub_str)

        # 解析销量
        sales = 0
        sales_val = get_val(COL_SALES)
        try:
            sales = int(float(str(sales_val))) if sales_val else 0
        except (ValueError, TypeError):
            pass

        return {
            "rank": row_idx - 1,  # 数据行号
            "image": str(get_val(COL_IMAGE) or ""),
            "title": str(title).strip(),
            "title_cn": "",  # 稍后翻译填充
            "url": str(get_val(COL_URL) or ""),
            "sku": str(get_val(COL_SKU) or ""),
            "brand": str(get_val(COL_BRAND) or ""),
            "category": str(get_val(COL_CATEGORY) or ""),
            "price_rub": price_rub,
            "price_rub_str": price_rub_str,
            "price_cny": round(price_rub * RUB_TO_CNY, 2),
            "sales": sales,
            "margin": str(get_val(COL_MARGIN) or ""),
            "weight": str(get_val(COL_WEIGHT) or ""),
            "volume": str(get_val(COL_VOLUME) or ""),
        }

    @staticmethod
    def _parse_rub_price(price_str: str) -> float:
        """解析卢布价格，如 '5312₽' → 5312.0"""
        if not price_str:
            return 0.0
        cleaned = price_str.replace("₽", "").replace(" ", "").replace(",", ".").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0.0

    # ================================================================
    # Excel 输出
    # ================================================================

    def _write_excel(self, matches: List[ProductMatch], products: List[dict]):
        """将匹配结果写入Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "1688货源匹配"

        # 定义输出列
        output_columns = [
            ("排名", 8),
            ("Ozon标题(俄语)", 45),
            ("中文关键词", 30),
            ("Ozon售价(₽)", 14),
            ("Ozon售价(¥)", 14),
            ("Ozon销量", 10),
            ("Ozon毛利率", 12),
            ("品牌", 12),
            ("类目", 20),
            ("1688商品标题", 45),
            ("1688批发价(¥)", 16),
            ("起批量", 10),
            ("1688链接", 35),
            ("供应商", 20),
            ("1688销量", 10),
            ("匹配方式", 12),
            ("相似度", 10),
            ("匹配评分", 10),
            ("预估毛利率", 12),
        ]

        # 样式定义
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_font = Font(name="微软雅黑", size=10)
        cell_alignment = Alignment(vertical="center", wrap_text=True)
        link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        # 图片匹配绿色背景，关键词匹配浅蓝背景
        image_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        keyword_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        both_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        # 写表头
        for col_idx, (col_name, col_width) in enumerate(output_columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # 写数据
        current_row = 2
        for match in matches:
            for m in match.matches:
                # 匹配方式颜色
                if m.search_type == "image":
                    row_fill = image_fill
                elif m.search_type == "both":
                    row_fill = both_fill
                else:
                    row_fill = keyword_fill

                values = [
                    match.ozon_rank,
                    match.ozon_title,
                    match.ozon_title_cn,
                    match.ozon_price_rub,
                    match.ozon_price_cny,
                    match.ozon_sales,
                    match.ozon_margin,
                    match.brand,
                    match.category,
                    m.title_1688,
                    m.display_price,
                    m.min_order,
                    m.detail_url,
                    m.shop_name,
                    "",  # 1688销量 - 暂不展示
                    m.search_type if m.search_type != "both" else "图片+关键词",
                    f"{m.similarity:.0%}" if m.similarity > 0 else "-",
                    m.score,
                    self.matcher.estimate_profit_margin(
                        match.ozon_price_cny, m.price_min
                    ),
                ]

                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    # 1688链接列设为超链接
                    if col_idx == 13 and isinstance(val, str) and val.startswith("http"):
                        cell.font = link_font
                        cell.hyperlink = val

                current_row += 1

            # 如果没有匹配，输出原始Ozon信息
            if not match.matches:
                values = [
                    match.ozon_rank,
                    match.ozon_title,
                    match.ozon_title_cn,
                    match.ozon_price_rub,
                    match.ozon_price_cny,
                    match.ozon_sales,
                    match.ozon_margin,
                    match.brand,
                    match.category,
                    "（未找到匹配）", "", "", "", "", "", "", "", "", "",
                ]
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                current_row += 1

        # 冻结首行
        ws.freeze_panes = "A2"

        # 添加筛选
        ws.auto_filter.ref = f"A1:{get_column_letter(len(output_columns))}{current_row - 1}"

        # 保存
        self.output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_file)
        wb.close()
        logger.info(f"结果已保存: {self.output_file} (共 {current_row - 2} 行数据)")
