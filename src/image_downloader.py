"""
Step 1: 从 Excel 批量下载 Ozon 商品图片到本地
"""
import os
import hashlib
import logging
from pathlib import Path
from typing import List, Dict
from urllib.parse import urlparse

import requests
import openpyxl
from tqdm import tqdm

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
IMAGE_DIR = PROJECT_ROOT / "output" / "images"
IMAGE_DIR.mkdir(parents=True, exist_ok=True)


def download_images(
    input_excel: str,
    limit: int = 0
) -> List[Dict]:
    """
    读取 Excel，下载所有商品图片到本地

    Args:
        input_excel: Excel 文件路径
        limit: 限制下载数量（0=全部）

    Returns:
        商品列表 [{row, title_ru, title_cn, image_local, category, price_rub, ...}, ...]
    """
    # 加载翻译字典
    translations = _load_translations()

    wb = openpyxl.load_workbook(input_excel, data_only=True)
    ws = wb.active
    logger.info(f"读取 Excel: {ws.max_row - 1} 条数据")

    products = []

    for row_idx in range(2, ws.max_row + 1):
        image_url = str(ws.cell(row=row_idx, column=2).value or "").strip()
        title_ru = str(ws.cell(row=row_idx, column=4).value or "").strip()
        detail_url = str(ws.cell(row=row_idx, column=5).value or "").strip()
        sku = str(ws.cell(row=row_idx, column=6).value or "").strip()
        brand = str(ws.cell(row=row_idx, column=7).value or "").strip()
        category = str(ws.cell(row=row_idx, column=8).value or "").strip()
        price_rub = str(ws.cell(row=row_idx, column=10).value or "").strip()
        sales = str(ws.cell(row=row_idx, column=11).value or "").strip()

        if not title_ru:
            continue

        title_cn = translations.get(title_ru, "")

        product = {
            "row": row_idx - 1,  # 行号（1-based，不含表头）
            "image_url": image_url,
            "image_local": "",
            "title_ru": title_ru,
            "title_cn": title_cn,
            "detail_url": detail_url,
            "sku": sku,
            "brand": brand,
            "category": category,
            "price_rub": price_rub,
            "sales": sales,
        }
        products.append(product)

        if limit and len(products) >= limit:
            break

    wb.close()

    # 下载图片
    logger.info(f"开始下载 {len(products)} 张图片...")
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        "Referer": "https://www.ozon.ru/",
    })

    success = 0
    for p in tqdm(products, desc="下载图片"):
        local_path = _download_single(session, p["image_url"])
        if local_path:
            p["image_local"] = str(local_path)
            success += 1

    logger.info(f"图片下载完成: {success}/{len(products)}")
    return products


def _load_translations() -> Dict[str, str]:
    """加载翻译 CSV"""
    import csv
    csv_path = PROJECT_ROOT / "output" / "titles_for_translation.csv"
    translations = {}
    if csv_path.exists():
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if len(row) >= 2 and row[0].strip() and row[1].strip():
                    translations[row[0].strip()] = row[1].strip()
    return translations


def _download_single(session: requests.Session, image_url: str) -> "Path | None":
    """下载单张图片"""
    if not image_url:
        return None

    # 生成唯一文件名
    name_hash = hashlib.md5(image_url.encode()).hexdigest()[:12]
    # 保留原始扩展名
    parsed = urlparse(image_url)
    ext = os.path.splitext(parsed.path)[1] or ".jpg"
    local_path = IMAGE_DIR / f"{name_hash}{ext}"

    # 已存在且有效则跳过
    if local_path.exists() and local_path.stat().st_size > 100:
        return local_path

    try:
        resp = session.get(image_url, timeout=15, stream=True)
        resp.raise_for_status()
        with open(local_path, "wb") as f:
            for chunk in resp.iter_content(8192):
                f.write(chunk)
        return local_path
    except Exception as e:
        logger.debug(f"下载失败 {image_url[:60]}: {e}")
        return None


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    products = download_images(
        str(PROJECT_ROOT / "data" / "Seerfar-Product20260614_200.xlsx"),
        limit=5,
    )
    for p in products:
        print(f"  [{p['row']}] {p['title_ru'][:50]} → 图片: {p['image_local']}")
