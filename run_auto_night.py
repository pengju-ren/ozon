#!/usr/bin/env python3
"""
🌙 全自动夜间跑机 — 无需任何手动操作
1. 复制你 Chrome 的登录态（1688已登录的Cookie）
2. 启动独立 Chrome 实例（带调试端口）
3. 逐张图片上 1688 以图搜款
4. 结果存 Excel，随时可查看进度

用法:
    python run_auto_night.py                # 跑全部200个
    python run_auto_night.py --limit 10     # 测试10个
"""
import sys
import os
import re
import csv
import json
import time
import random
import shutil
import hashlib
import asyncio
import logging
import subprocess
from pathlib import Path
from typing import List, Dict, Optional

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(ROOT / "output" / "auto_night.log", mode="a", encoding="utf-8"),
    ],
)
logger = logging.getLogger("auto")

DEBUG_PORT = 9222
CDP_URL = f"http://localhost:{DEBUG_PORT}"
IMAGE_DIR = ROOT / "output" / "images"
CHECKPOINT = ROOT / "output" / "auto_checkpoint.json"
OUTPUT_EXCEL = ROOT / "output" / "1688_auto_result.xlsx"
CHROME_REAL_PROFILE = Path.home() / "Library/Application Support/Google/Chrome"


# ============================================================
# Step 1: 复用 Chrome 登录态
# ============================================================
def setup_chrome_profile() -> Path:
    """
    复制用户的 Chrome 配置（含 1688 登录 Cookie）到临时目录
    这样启动的新 Chrome 直接就是登录状态
    """
    temp_profile = ROOT / "output" / "chrome_night_profile"

    # 如果之前跑过，直接复用
    if (temp_profile / "Cookies").exists():
        logger.info("复用已有 Chrome Profile")
        return temp_profile

    logger.info("正在复制 Chrome 登录态...")

    # 杀掉已运行的 Chrome（解除文件锁）
    subprocess.run(["killall", "Google Chrome"], capture_output=True)
    time.sleep(2)

    # 只复制关键的 Cookie 和登录相关文件
    temp_profile.mkdir(parents=True, exist_ok=True)
    key_files = ["Cookies", "Cookies-journal", "Login Data", "Login Data-journal",
                 "Preferences", "Network", "Local Storage", "Session Storage"]
    for item in key_files:
        src = CHROME_REAL_PROFILE / item
        dst = temp_profile / item
        if src.exists():
            try:
                if src.is_dir():
                    shutil.copytree(src, dst, dirs_exist_ok=True)
                else:
                    shutil.copy2(src, dst)
            except Exception:
                pass

    # 也复制 Default/ 下的
    default_src = CHROME_REAL_PROFILE / "Default"
    default_dst = temp_profile / "Default"
    if default_src.exists():
        os.makedirs(default_dst, exist_ok=True)
        for item in key_files:
            s = default_src / item
            d = default_dst / item
            if s.exists():
                try:
                    if s.is_dir():
                        shutil.copytree(s, d, dirs_exist_ok=True)
                    else:
                        shutil.copy2(s, d)
                except Exception:
                    pass

    logger.info("✅ Chrome Profile 就绪")
    return temp_profile


def launch_chrome_with_profile(profile_dir: Path) -> Optional[subprocess.Popen]:
    """用已有 Profile 启动 Chrome + 开启调试端口"""
    chrome = None
    for p in [
        Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
        Path.home() / "Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
    ]:
        if p.exists():
            chrome = p
            break

    if not chrome:
        logger.error("找不到 Chrome")
        return None

    logger.info("启动 Chrome（登录态已复用）...")
    proc = subprocess.Popen(
        [str(chrome),
         f"--remote-debugging-port={DEBUG_PORT}",
         f"--user-data-dir={profile_dir}",
         "--no-first-run", "--no-default-browser-check",
         "--disable-extensions", "--disable-sync",
         ],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        start_new_session=True,
    )

    # 等 Chrome 就绪
    for _ in range(20):
        time.sleep(1)
        try:
            import urllib.request
            urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2)
            logger.info("✅ Chrome 就绪")
            return proc
        except Exception:
            pass
    return None


# ============================================================
# Step 2: 数据
# ============================================================
def load_data(limit: int = 200) -> List[Dict]:
    import openpyxl, requests as req

    translations = {}
    csv_path = ROOT / "output" / "titles_for_translation.csv"
    if csv_path.exists():
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            for r in csv.reader(f):
                if len(r) >= 2 and r[0].strip() and r[1].strip():
                    translations[r[0].strip()] = r[1].strip()

    def lookup(text):
        if text in translations: return translations[text]
        for k, v in translations.items():
            if k.startswith(text) or text.startswith(k): return v
        return ""

    wb = openpyxl.load_workbook(
        ROOT / "data" / "Seerfar-Product20260614_200.xlsx", data_only=True)
    ws = wb.active
    products = []
    for r in range(2, ws.max_row + 1):
        title = str(ws.cell(row=r, column=4).value or "").strip()
        if not title: continue
        products.append({
            "row": r - 1,
            "image_url": str(ws.cell(row=r, column=2).value or "").strip(),
            "title_ru": title,
            "title_cn": lookup(title),
            "brand": str(ws.cell(row=r, column=7).value or "").strip(),
            "category": str(ws.cell(row=r, column=8).value or "").strip(),
            "price_rub": str(ws.cell(row=r, column=10).value or "").strip(),
            "sales": str(ws.cell(row=r, column=11).value or "").strip(),
        })
        if limit and len(products) >= limit: break
    wb.close()

    session = req.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
        "Referer": "https://www.ozon.ru/",
    })
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    got = 0
    for p in products:
        if not p["image_url"]: continue
        name = hashlib.md5(p["image_url"].encode()).hexdigest()[:12] + ".jpg"
        path = IMAGE_DIR / name
        if path.exists() and path.stat().st_size > 100:
            p["image_local"] = str(path); got += 1; continue
        try:
            resp = session.get(p["image_url"], timeout=15)
            path.write_bytes(resp.content)
            p["image_local"] = str(path); got += 1
        except Exception: pass
    logger.info(f"数据就绪: {len(products)} 商品 | 图片 {got} 张")
    return products


# ============================================================
# Step 3: CDP 搜索器
# ============================================================
class NightSearcher:
    def __init__(self):
        self._pw = None
        self._browser = None
        self._page = None

    async def connect(self):
        from playwright.async_api import async_playwright
        self._pw = await async_playwright().start()
        self._browser = await self._pw.chromium.connect_over_cdp(CDP_URL)
        pages = self._browser.contexts[0].pages if self._browser.contexts else []
        self._page = pages[0] if pages else await self._browser.contexts[0].new_page()
        logger.info(f"已连接 Chrome: {self._page.url[:60]}")

    async def search(self, product: Dict) -> Dict:
        result = {"status": "no_result", "products": []}
        img_path = product.get("image_local", "")
        if not img_path or not Path(img_path).exists():
            result["status"] = "no_image"; return result

        try:
            # 打开 1688
            await self._page.goto(
                "https://www.1688.com/",
                wait_until="domcontentloaded", timeout=20000)
            await _sleep(1.5, 3)

            # 检查登录
            if "login" in self._page.url.lower():
                result["status"] = "not_logged_in"
                logger.warning("1688 未登录 — Cookie 可能已过期")
                return result

            # 点相机
            clicked = False
            for sel in [
                '.search-img-btn', 'img[alt*="拍照"]', 'img[alt*="图片"]',
                '.alisearch-camera', 'i[class*="camera"]',
                '[class*="image-search"]', 'span[class*="camera"]',
            ]:
                btn = await self._page.query_selector(sel)
                if btn:
                    try:
                        await btn.click()
                        await _sleep(0.5, 1.5)
                        clicked = True
                        break
                    except Exception:
                        continue

            if not clicked:
                await self._page.goto(
                    "https://s.1688.com/selloffer/offer_search.htm?imageSearch=1",
                    wait_until="domcontentloaded", timeout=15000)
                await _sleep(1, 2)

            # 上传图片
            file_input = None
            for _ in range(3):
                await _sleep(0.5, 1)
                file_input = await self._page.query_selector('input[type="file"]')
                if file_input: break

            if not file_input:
                result["status"] = "no_upload"; return result

            await file_input.set_input_files(img_path)
            await _sleep(3.5, 6)

            # 解析
            products = await self._extract()
            if products:
                result["status"] = "success"
                result["products"] = products[:12]
        except Exception as e:
            logger.debug(f"搜索异常: {e}")
            result["status"] = "error"
        return result

    async def _extract(self) -> List[Dict]:
        await _sleep(1, 2)
        html = await self._page.content()
        results = []

        # data 属性
        oids = re.findall(r'data-offerid="(\d+)"', html)
        titles = re.findall(r'data-offer-title="([^"]*)"', html)
        prices = re.findall(r'data-offer-price="([^"]*)"', html)

        if oids:
            for i in range(min(len(oids), 12)):
                p = {
                    "offer_id": oids[i],
                    "title": titles[i] if i < len(titles) else "",
                    "price_text": prices[i] if i < len(prices) else "",
                    "detail_url": f"https://detail.1688.com/offer/{oids[i]}.html",
                }
                lo, hi = _pp(p["price_text"]); p["price_low"] = lo; p["price_high"] = hi
                results.append(p)

        # CSS 卡片
        if not results:
            for sel in [
                '.space-offer-card', '.sm-offer-item',
                'div[class*="offer-item"]', 'li[class*="offer"]',
            ]:
                cards = await self._page.query_selector_all(sel)
                if cards:
                    for card in cards[:12]:
                        try:
                            t_el = await card.query_selector('.title, h3 a, [class*="subject"]')
                            title = (await t_el.inner_text()).strip() if t_el else ""
                            pr_el = await card.query_selector('[class*="price"]')
                            price = (await pr_el.inner_text()).strip() if pr_el else ""
                            l_el = await card.query_selector('a[href*="offer"]')
                            href = await l_el.get_attribute("href") if l_el else ""
                            if href and href.startswith("//"): href = "https:" + href
                            s_el = await card.query_selector('[class*="seller"], [class*="shop"], [class*="company"]')
                            shop = (await s_el.inner_text()).strip() if s_el else ""
                            q_el = await card.query_selector('[class*="quantity"], [class*="min"]')
                            moq = (await q_el.inner_text()).strip() if q_el else ""
                            lo, hi = _pp(price)
                            results.append({
                                "title": title, "price_text": price,
                                "price_low": lo, "price_high": hi,
                                "detail_url": href, "shop": shop, "min_order": moq,
                            })
                        except Exception:
                            continue
                    break

        return results

    async def close(self):
        if self._pw: await self._pw.stop()


def _pp(text):  # parse price
    if not text: return 0.0, 0.0
    text = str(text).replace("¥", "").replace("￥", "").replace(",", "").strip()
    for sep in ("-", "~", "—", " "):
        if sep in text:
            p = text.split(sep)
            try:
                a, b = float(p[0].strip()), float(p[-1].strip())
                return min(a, b), max(a, b)
            except ValueError: continue
    try: v = float(text); return v, v
    except ValueError: return 0.0, 0.0


async def _sleep(lo, hi):
    await asyncio.sleep(random.uniform(lo, hi))


# ============================================================
# Step 4: 主循环
# ============================================================
async def run(products: List[Dict]):
    done = set()
    if CHECKPOINT.exists():
        with open(CHECKPOINT) as f:
            done = {r["row"] for r in json.load(f) if r.get("status") == "success"}
        logger.info(f"断点: 已完成 {len(done)}")

    searcher = NightSearcher()
    await searcher.connect()

    # 检测登录
    await searcher._page.goto("https://www.1688.com", wait_until="domcontentloaded", timeout=15000)
    await asyncio.sleep(2)
    if "login" in searcher._page.url.lower():
        logger.error("❌ 未登录！可能是 Cookie 过期，请手动登录后重新运行")
        await searcher.close()
        return []

    logger.info("✅ 1688 登录有效")

    results = []
    ok = fail = 0
    try:
        for i, p in enumerate(products):
            if p["row"] in done: continue
            logger.info(f"\n── [{i+1}/{len(products)}] #{p['row']} {p['title_ru'][:45]}")
            r = {"row": p["row"], "ozon_title": p["title_ru"],
                 "ozon_title_cn": p["title_cn"], "ozon_price_rub": p["price_rub"],
                 "ozon_image": p["image_url"], "category": p["category"],
                 "brand": p["brand"], "sales": p["sales"],
                 "status": "", "products": []}
            sr = await searcher.search(p)
            r["status"] = sr["status"]
            r["products"] = sr["products"]
            if sr["status"] == "success": ok += 1
            else: fail += 1
            results.append(r)

            CHECKPOINT.parent.mkdir(parents=True, exist_ok=True)
            with open(CHECKPOINT, "w") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)

            logger.info(f"  📊 成功:{ok} 失败:{fail}")

            # 如检测到登出，停止
            if sr["status"] == "not_logged_in":
                logger.error("Cookie 失效，停止")
                break

            await _sleep(5, 12)
    finally:
        await searcher.close()
    return results


# ============================================================
# Step 5: 存 Excel
# ============================================================
def save_excel(results):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "1688以图搜款"
    headers = ["序号", "Ozon标题", "中文关键词", "Ozon售价₽", "类目", "品牌",
               "1688标题", "1688批发价", "起批量", "店铺", "1688链接", "状态"]
    widths = [6, 35, 22, 10, 16, 10, 35, 14, 8, 14, 30, 8]
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    cf = Font(name="微软雅黑", size=10)
    lf = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    bd = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    gf = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    rf = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hf
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = bd
        ws.column_dimensions[get_column_letter(ci)].width = w

    ri = 2
    for r in results:
        base = [r["row"], r["ozon_title"], r["ozon_title_cn"],
                r["ozon_price_rub"], r["category"], r["brand"]]
        if r["products"]:
            for p in r["products"]:
                vals = base + [p.get("title", ""),
                    p.get("price_text", "") or f"{p.get('price_low', 0)}-{p.get('price_high', 0)}",
                    p.get("min_order", ""), p.get("shop", ""),
                    p.get("detail_url", ""), r["status"]]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=ri, column=ci, value=v)
                    c.font = cf; c.border = bd; c.fill = gf
                    if ci == 11 and str(v).startswith("http"):
                        c.font = lf; c.hyperlink = str(v)
                ri += 1
        else:
            vals = base + ["", "", "", "", "", r["status"]]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = cf; c.border = bd; c.fill = rf
            ri += 1
    ws.freeze_panes = "A2"
    wb.save(str(OUTPUT_EXCEL)); wb.close()
    logger.info(f"📊 结果: {OUTPUT_EXCEL}")


# ============================================================
# main
# ============================================================
async def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=200)
    args = ap.parse_args()

    logger.info("=" * 50)
    logger.info("🌙 全自动夜间跑机启动")
    logger.info("=" * 50)

    # 1. 准备 Chrome Profile（复制登录态）
    profile = setup_chrome_profile()

    # 2. 数据
    products = load_data(limit=args.limit)
    if not products:
        logger.error("无数据"); return

    # 3. 启动 Chrome
    chrome_proc = launch_chrome_with_profile(profile)
    if not chrome_proc:
        logger.error("Chrome 启动失败"); return

    # 4. 跑流程
    try:
        results = await run(products)
        if results:
            save_excel(results)
            ok = sum(1 for r in results if r["status"] == "success")
            logger.info(f"\n✅ 完成! {ok}/{len(results)} 个成功")
    finally:
        if chrome_proc:
            chrome_proc.terminate()
            logger.info("Chrome 已关闭")


if __name__ == "__main__":
    asyncio.run(main())
