#!/usr/bin/env python3
"""
🚀 CDP 模式 — 接管真实 Chrome 浏览器以图搜款
原理：跟影刀 RPA 一样，用你电脑上已登录 1688 的真实 Chrome

使用步骤:
  1. 关闭所有 Chrome 窗口
  2. 运行: python run_cdp_search.py
  3. 脚本自动启动 Chrome，你在里面登录 1688
  4. 登录后按 Enter，开始全自动跑

特点:
  - 使用你真实的 Chrome 浏览器（Cookie/登录态都在）
  - 1688 看到的就是你的正常浏览器，不会封
  - 模拟鼠标点击、文件上传
  - 每张图间隔 8-15 秒随机，像真人操作
"""
import sys
import os
import re
import csv
import json
import time
import random
import shutil
import hashlib
import asyncio
import logging
import subprocess
import webbrowser
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass, field, asdict

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

# ---- 日志 ----
LOG_FILE = ROOT / "output" / "cdp_search.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_FILE, mode="a", encoding="utf-8"),
    ],
)
logger = logging.getLogger("cdp")

IMAGE_DIR = ROOT / "output" / "images"
CHECKPOINT = ROOT / "output" / "cdp_checkpoint.json"
OUTPUT_EXCEL = ROOT / "output" / "1688_cdp_result.xlsx"
DEBUG_PORT = 9222
CDP_URL = f"http://localhost:{DEBUG_PORT}"


# ============================================================
# Step 1: 启动 Chrome（带调试端口）
# ============================================================
def find_chrome() -> Optional[Path]:
    """查找 Chrome 可执行文件"""
    paths = [
        Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
        Path.home() / "Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
    ]
    for p in paths:
        if p.exists():
            return p
    return None


def launch_chrome() -> Optional[subprocess.Popen]:
    """
    启动 Chrome 并开启远程调试端口
    使用独立的临时 profile（不影响你的正常 Chrome 数据）
    """
    chrome = find_chrome()
    if not chrome:
        logger.error("找不到 Google Chrome！请安装 Chrome 浏览器")
        return None

    # 使用临时用户目录
    profile_dir = ROOT / "output" / "chrome_cdp_profile"
    if profile_dir.exists():
        shutil.rmtree(profile_dir, ignore_errors=True)
    profile_dir.mkdir(parents=True, exist_ok=True)

    logger.info("正在启动 Chrome（带调试端口）...")

    proc = subprocess.Popen(
        [
            str(chrome),
            f"--remote-debugging-port={DEBUG_PORT}",
            f"--user-data-dir={profile_dir}",
            "--no-first-run",
            "--no-default-browser-check",
            "--disable-extensions",
        ],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        start_new_session=True,
    )

    # 等待 Chrome 启动
    for _ in range(15):
        time.sleep(1)
        try:
            import urllib.request
            urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2)
            logger.info("✅ Chrome 已启动，调试端口就绪")
            return proc
        except Exception:
            pass

    logger.error("Chrome 启动失败或超时")
    return None


def wait_for_1688_login():
    """让用户手动登录 1688"""
    logger.info("\n" + "=" * 60)
    logger.info("📌 请在 Chrome 中完成以下操作：")
    logger.info("   1. 打开 https://www.1688.com")
    logger.info("   2. 点击右上角「登录」")
    logger.info("   3. 用淘宝/支付宝扫码登录")
    logger.info("   4. 确认登录成功后，回到终端按 Enter 键")
    logger.info("=" * 60)

    input("\n>>> 登录完成后按 Enter 继续...")


# ============================================================
# Step 2: 数据准备
# ============================================================
def load_data(limit: int = 200) -> List[Dict]:
    """加载商品数据 + 翻译 + 下载图片"""
    import openpyxl
    import requests as req

    # 加载翻译
    translations = {}
    csv_path = ROOT / "output" / "titles_for_translation.csv"
    if csv_path.exists():
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            for r in csv.reader(f):
                if len(r) >= 2 and r[0].strip() and r[1].strip():
                    translations[r[0].strip()] = r[1].strip()

    def lookup(text: str) -> str:
        if text in translations:
            return translations[text]
        for k, v in translations.items():
            if k.startswith(text) or text.startswith(k):
                return v
        return ""

    wb = openpyxl.load_workbook(
        ROOT / "data" / "Seerfar-Product20260614_200.xlsx", data_only=True
    )
    ws = wb.active

    products = []
    for r in range(2, ws.max_row + 1):
        title = str(ws.cell(row=r, column=4).value or "").strip()
        if not title:
            continue
        products.append({
            "row": r - 1,
            "image_url": str(ws.cell(row=r, column=2).value or "").strip(),
            "title_ru": title,
            "title_cn": lookup(title),
            "brand": str(ws.cell(row=r, column=7).value or "").strip(),
            "category": str(ws.cell(row=r, column=8).value or "").strip(),
            "price_rub": str(ws.cell(row=r, column=10).value or "").strip(),
            "sales": str(ws.cell(row=r, column=11).value or "").strip(),
        })
        if limit and len(products) >= limit:
            break
    wb.close()

    # 下载图片
    logger.info(f"准备 {len(products)} 个商品，下载图片...")
    session = req.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
        "Referer": "https://www.ozon.ru/",
    })
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    got = 0
    for p in products:
        if not p["image_url"]:
            continue
        name = hashlib.md5(p["image_url"].encode()).hexdigest()[:12] + ".jpg"
        path = IMAGE_DIR / name
        if path.exists() and path.stat().st_size > 100:
            p["image_local"] = str(path)
            got += 1
            continue
        try:
            resp = session.get(p["image_url"], timeout=15)
            resp.raise_for_status()
            path.write_bytes(resp.content)
            p["image_local"] = str(path)
            got += 1
        except Exception:
            pass
    logger.info(f"图片就绪: {got}/{len(products)}")
    return products


# ============================================================
# Step 3: CDP 以图搜款
# ============================================================
class CDPImageSearcher:
    """通过 CDP 接管真实 Chrome，模拟真人以图搜款"""

    def __init__(self):
        self._playwright = None
        self._browser = None
        self._page = None
        self.success_count = 0
        self.fail_count = 0

    async def connect(self):
        """连接到已有 Chrome 实例"""
        from playwright.async_api import async_playwright

        self._playwright = await async_playwright().start()

        try:
            self._browser = await self._playwright.chromium.connect_over_cdp(CDP_URL)
        except Exception as e:
            logger.error(f"连接 Chrome 失败: {e}")
            logger.error(f"请确认 Chrome 以调试模式运行: {CDP_URL}")
            raise

        # 获取或创建页面
        pages = self._browser.contexts[0].pages if self._browser.contexts else []
        if pages:
            self._page = pages[0]
        else:
            self._page = await self._browser.contexts[0].new_page()

        logger.info("✅ 已连接到你的 Chrome 浏览器")
        logger.info(f"   当前页面: {self._page.url[:60]}")

    async def search_one_product(self, product: Dict) -> Dict:
        """
        对一个商品执行以图搜款
        返回: {"status": "success/no_result/error", "products": [...]}
        """
        img_path = product.get("image_local", "")
        result = {
            "status": "no_result",
            "products": [],
        }

        if not img_path or not Path(img_path).exists():
            result["status"] = "no_image"
            return result

        try:
            # ---- 打开 1688 首页 ----
            await self._page.goto(
                "https://www.1688.com/",
                wait_until="domcontentloaded",
                timeout=20000,
            )
            await self._random_delay(1.5, 3.0)

            # 检查是否在登录页（如果没登录会被重定向）
            if "login" in self._page.url.lower():
                logger.warning("⚠️ 1688 未登录！请先登录后再运行")
                result["status"] = "not_logged_in"
                return result

            # ---- 点击以图搜货的相机按钮 ----
            camera_clicked = False
            selectors = [
                '.search-img-btn',
                'img[alt*="拍照搜"]',
                'img[alt*="以图"]',
                'img[alt*="图片"]',
                '.alisearch-camera',
                'span[class*="icon-camera"]',
                'i[class*="ali-icon-camera"]',
                'div[class*="search-camera"]',
                'img[src*="camera"]',
            ]

            for sel in selectors:
                btn = await self._page.query_selector(sel)
                if btn:
                    try:
                        await btn.click()
                        await self._random_delay(0.5, 1.5)
                        camera_clicked = True
                        logger.debug(f"  点击相机：{sel}")
                        break
                    except Exception:
                        continue

            if not camera_clicked:
                logger.debug("  没找到相机按钮，尝试其他方式...")

            # ---- 上传图片 ----
            file_input = None
            for _ in range(3):  # 等弹窗出现
                await self._random_delay(0.5, 1)

                # 主选择器：file input
                file_input = await self._page.query_selector('input[type="file"]')
                if file_input:
                    break

                # 备用：找任何包含 upload 的元素
                upload_btns = await self._page.query_selector_all(
                    'input[accept*="image"], '
                    '[class*="upload-img"], '
                    'input[id*="upload"], '
                    'form[enctype*="multipart"] input'
                )
                if upload_btns:
                    file_input = upload_btns[0]
                    break

            if not file_input:
                logger.warning("  找不到上传按钮")
                result["status"] = "no_upload_btn"
                return result

            # 上传
            await file_input.set_input_files(img_path)
            logger.debug(f"  已上传: {Path(img_path).name}")
            await self._random_delay(3.5, 6.0)  # 等识别

            # ---- 解析结果 ----
            products = await self._extract_products()
            if products:
                result["status"] = "success"
                result["products"] = products[:12]  # 前12个
                logger.info(f"  ✅ 找到 {len(products)} 个商品")
            else:
                logger.info("  ⚠️ 无匹配结果")

        except Exception as e:
            logger.error(f"  搜索异常: {e}")
            result["status"] = "error"

        return result

    async def _extract_products(self) -> List[Dict]:
        """从当前搜索结果页提取商品列表"""
        products = []

        try:
            await self._random_delay(1, 2)
            html = await self._page.content()

            # 检查特殊状态
            if "暂无相关" in html or "没找到" in html:
                return []
            if "captcha" in html.lower():
                logger.warning("  ⚠️ 遇到验证码！")
                return []

            # 策略 1: data 属性
            offer_ids = re.findall(r'data-offerid="(\d+)"', html)
            titles = re.findall(r'data-offer-title="([^"]*)"', html)
            prices = re.findall(r'data-offer-price="([^"]*)"', html)

            if offer_ids:
                for i in range(min(len(offer_ids), 12)):
                    products.append({
                        "offer_id": offer_ids[i],
                        "title": titles[i] if i < len(titles) else "",
                        "price_text": prices[i] if i < len(prices) else "",
                        "detail_url": f"https://detail.1688.com/offer/{offer_ids[i]}.html",
                    })

            # 策略 2: CSS 卡片
            if not products:
                card_selectors = [
                    '.space-offer-card',
                    '.sm-offer-item',
                    'div[class*="offer-item"]',
                    'li[class*="offer"]',
                ]
                for sel in card_selectors:
                    cards = await self._page.query_selector_all(sel)
                    if cards:
                        for card in cards[:12]:
                            p = await self._parse_card(card)
                            if p:
                                products.append(p)
                        break

            # 策略 3: 正则暴力提取
            if not products:
                price_matches = re.findall(r'[¥￥]\s*([\d,.]+)', html)
                link_matches = re.findall(
                    r'href="(https?://detail\.1688\.com/offer/\d+\.html)"', html
                )
                for i, link in enumerate(link_matches[:12]):
                    products.append({
                        "title": f"匹配商品{i+1}",
                        "price_text": price_matches[i] if i < len(price_matches) else "",
                        "detail_url": link,
                    })

        except Exception as e:
            logger.debug(f"  解析异常: {e}")

        # 计算价格
        for p in products:
            price_text = p.get("price_text", "")
            lo, hi = _parse_price(price_text)
            p["price_low"] = lo
            p["price_high"] = hi

        return products

    async def _parse_card(self, card) -> Optional[Dict]:
        try:
            title_el = await card.query_selector(
                '.title, h3 a, [class*="subject"], .offer-title'
            )
            title = (await title_el.inner_text()).strip() if title_el else ""

            price_el = await card.query_selector(
                '.price, [class*="price-num"], [class*="price-text"]'
            )
            price_text = (await price_el.inner_text()).strip() if price_el else ""

            link_el = await card.query_selector('a[href*="offer"]')
            href = await link_el.get_attribute("href") if link_el else ""
            if href and href.startswith("//"):
                href = f"https:{href}"

            shop_el = await card.query_selector(
                '[class*="seller"], [class*="shop"], [class*="company"]'
            )
            shop = (await shop_el.inner_text()).strip() if shop_el else ""

            moq_el = await card.query_selector('[class*="quantity"], [class*="min"]')
            moq = (await moq_el.inner_text()).strip() if moq_el else ""

            return {
                "title": title,
                "price_text": price_text,
                "detail_url": href,
                "shop": shop,
                "min_order": moq,
            }
        except Exception:
            return None

    async def close(self):
        if self._playwright:
            await self._playwright.stop()

    @staticmethod
    async def _random_delay(lo: float, hi: float):
        await asyncio.sleep(random.uniform(lo, hi))


def _parse_price(text: str) -> tuple:
    if not text:
        return 0.0, 0.0
    text = str(text).replace("¥", "").replace("￥", "").replace(",", "").strip()
    for sep in ("-", "~", "—", " "):
        if sep in text:
            parts = text.split(sep)
            try:
                a, b = float(parts[0].strip()), float(parts[-1].strip())
                return min(a, b), max(a, b)
            except ValueError:
                continue
    try:
        p = float(text)
        return p, p
    except ValueError:
        return 0.0, 0.0


# ============================================================
# Step 4: 主循环
# ============================================================
async def run_pipeline(limit: int = 200):
    """主流程"""
    logger.info("📦 加载数据...")
    products = load_data(limit=limit)

    # 断点恢复
    done_ids = set()
    if CHECKPOINT.exists():
        with open(CHECKPOINT) as f:
            old = json.load(f)
            done_ids = {r["row"] for r in old if r.get("status") == "success"}
            logger.info(f"断点恢复: 已完成 {len(done_ids)} 个")

    # 连接 Chrome
    searcher = CDPImageSearcher()
    await searcher.connect()

    # 检查登录状态
    await searcher._page.goto("https://www.1688.com", wait_until="domcontentloaded", timeout=15000)
    await asyncio.sleep(2)
    if "login" in searcher._page.url.lower():
        logger.error("❌ 1688 未登录！请手动登录后重试")
        await searcher.close()
        return
    logger.info("✅ 1688 登录状态正常")

    results = []
    try:
        total = len(products)
        for i, product in enumerate(products):
            if product["row"] in done_ids:
                continue

            logger.info(f"\n{'─'*50}")
            logger.info(f"[{i+1}/{total}] #{product['row']} "
                        f"{product['title_ru'][:45]}")

            result = {
                "row": product["row"],
                "ozon_title": product["title_ru"],
                "ozon_title_cn": product["title_cn"],
                "ozon_price_rub": product["price_rub"],
                "ozon_image": product["image_url"],
                "category": product["category"],
                "brand": product["brand"],
                "sales": product["sales"],
                "status": "",
                "products": [],
            }

            r = await searcher.search_one_product(product)
            result["status"] = r["status"]
            result["products"] = r["products"]

            if r["status"] == "success":
                searcher.success_count += 1
            else:
                searcher.fail_count += 1

            results.append(result)

            # 保存检查点
            CHECKPOINT.parent.mkdir(parents=True, exist_ok=True)
            with open(CHECKPOINT, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)

            # 进度
            logger.info(f"  📊 累计: ✅{searcher.success_count} ⚠️{searcher.fail_count}")

            # 真人间隔
            delay = random.uniform(5.0, 12.0)
            logger.info(f"  ⏳ 等待 {delay:.0f}s...")
            await asyncio.sleep(delay)

    finally:
        await searcher.close()

    # 输出统计
    logger.info(f"\n{'='*60}")
    logger.info(f"🏁 完成! "
                f"成功:{searcher.success_count} | "
                f"失败:{searcher.fail_count}")
    return results


def save_excel(results: List[Dict]):
    """保存结果到 Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1688以图搜款-CDP"

    headers = [
        "序号", "Ozon标题", "中文关键词", "Ozon售价₽",
        "类目", "品牌", "Ozon销量",
        "1688商品标题", "1688批发价", "起批量",
        "店铺", "1688链接", "状态",
    ]
    widths = [6, 35, 22, 10, 16, 10, 8, 35, 14, 8, 14, 30, 8]

    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="微软雅黑", size=10)
    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    row_idx = 2
    for r in results:
        base = [
            r["row"], r["ozon_title"], r["ozon_title_cn"],
            r["ozon_price_rub"], r["category"], r["brand"], r["sales"],
        ]
        if r["products"]:
            for p in r["products"]:
                vals = base + [
                    p.get("title", ""),
                    p.get("price_text", "") or f"{p.get('price_low', 0)}-{p.get('price_high', 0)}",
                    p.get("min_order", ""),
                    p.get("shop", ""),
                    p.get("detail_url", ""),
                    r["status"],
                ]
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=ci, value=v)
                    cell.font = cell_font
                    cell.border = border
                    cell.fill = green_fill
                    if ci == 12 and str(v).startswith("http"):
                        cell.font = link_font
                        cell.hyperlink = str(v)
                row_idx += 1
        else:
            vals = base + ["", "", "", "", "", r["status"]]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.font = cell_font
                cell.border = border
                cell.fill = red_fill
            row_idx += 1

    ws.freeze_panes = "A2"
    wb.save(str(OUTPUT_EXCEL))
    wb.close()
    logger.info(f"📊 结果保存: {OUTPUT_EXCEL}")


# ============================================================
# 入口
# ============================================================
async def main():
    import argparse
    p = argparse.ArgumentParser(description="CDP 模式 1688 以图搜款")
    p.add_argument("--limit", type=int, default=200, help="商品数量(0=全部)")
    p.add_argument("--no-launch", action="store_true", help="不启动Chrome(手动启动)")
    args = p.parse_args()

    logger.info("=" * 60)
    logger.info("🚀 CDP 模式 — 接管真实 Chrome 以图搜款")
    logger.info("=" * 60)

    # 启动 Chrome
    chrome_proc = None
    if not args.no_launch:
        chrome_proc = launch_chrome()
        if not chrome_proc:
            return

        # 让用户登录
        wait_for_1688_login()

    # 跑流程
    results = await run_pipeline(limit=args.limit)

    # 保存
    if results:
        save_excel(results)

    # 清理
    if chrome_proc:
        chrome_proc.terminate()
        logger.info("Chrome 已关闭")


if __name__ == "__main__":
    asyncio.run(main())
