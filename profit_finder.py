#!/usr/bin/env python3
"""
1688 以图搜商 + 利润核算 — 稳定版

核心: 直连1688 H5 API (h5api.m.1688.com)，不使用浏览器
      图片→JPEG压缩→base64上传→搜索→解析→利润计算

用法:
  python profit_finder.py --limit 3              # 测试3条
  python profit_finder.py --limit 10 --no-image  # 仅关键词搜索
  python profit_finder.py --start 5 --limit 10   # 断点续跑
"""
import sys
import os
import re
import csv
import json
import time
import hashlib
import asyncio
import logging
import argparse
from pathlib import Path
from io import BytesIO
from typing import List, Dict, Optional
from dataclasses import dataclass, field, asdict

import aiohttp
import openpyxl
from yarl import URL
from PIL import Image

# 项目根
ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

# 复用现有签名函数
from search1688api.utils import generate_sign

# ============================================================
# 配置
# ============================================================
APP_KEY = "12574478"
BASE_URL = "https://h5api.m.1688.com/h5/mtop.relationrecommend.wirelessrecommend.recommend/2.0/"
RUB_TO_CNY = 0.078
DOMESTIC_SHIPPING = 5.0       # 国内运费估算(元)
OZON_COMMISSION_RATE = 0.10   # Ozon佣金率
IMAGE_DIR = ROOT / "output" / "images"
OUTPUT_DIR = ROOT / "output"
CHECKPOINT_FILE = ROOT / "output" / "profit_checkpoint.json"
LOG_FILE = ROOT / "output" / "profit_finder.log"
DEFAULT_INPUT = ROOT / "data" / "Seerfar-Product20260614_200.xlsx"
DEFAULT_OUTPUT = ROOT / "output" / "profit_result.xlsx"
CSV_TRANSLATION = ROOT / "output" / "titles_for_translation.csv"

IMAGE_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# 日志
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_FILE, mode="a", encoding="utf-8"),
    ],
)
logger = logging.getLogger("profit")


# ============================================================
# 数据结构
# ============================================================
@dataclass
class Product1688:
    """1688匹配商品"""
    offer_id: str = ""
    title: str = ""
    price_low: float = 0.0
    price_high: float = 0.0
    min_order: str = ""
    shop_name: str = ""
    location: str = ""
    detail_url: str = ""
    image_url: str = ""
    sales: int = 0
    similarity: float = 0.0

    @property
    def display_price(self) -> str:
        if self.price_low <= 0:
            return "-"
        if self.price_low == self.price_high:
            return f"{self.price_low:.2f}"
        return f"{self.price_low:.2f}-{self.price_high:.2f}"


@dataclass
class SearchResult:
    """一个Ozon商品的完整搜索结果"""
    row: int = 0
    ozon_title: str = ""
    ozon_title_cn: str = ""
    ozon_price_rub: str = ""
    ozon_sales: str = ""
    ozon_image_url: str = ""
    category: str = ""
    brand: str = ""
    image_local: str = ""
    status: str = "pending"     # success / no_results / no_image / error
    error_msg: str = ""
    products: List[dict] = field(default_factory=list)


# ============================================================
# H5 API 搜索器
# ============================================================
class H5Searcher:
    """直连1688 H5 API — 图片上传 + 以图搜款 + 关键词搜索"""

    def __init__(self):
        self._session: Optional[aiohttp.ClientSession] = None
        self._token_part: str = ""
        self._cookies: Dict[str, str] = {}

    # ---- 生命周期 ----
    async def start(self):
        self._session = aiohttp.ClientSession()
        await self._collect_cookies()
        await self._get_token()
        logger.info("H5Searcher 初始化完成")

    async def close(self):
        if self._session:
            await self._session.close()

    async def __aenter__(self):
        await self.start()
        return self

    async def __aexit__(self, *args):
        await self.close()

    # ---- 步骤1: 收集 Cookie ----
    async def _collect_cookies(self):
        """访问1688各页面收集cookie"""
        urls = [
            "https://www.1688.com",
            "https://s.1688.com",
            "https://login.1688.com",
            "https://s.1688.com/selloffer/offer_search.htm?keywords=sample",
        ]
        headers = {
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/142.0.0.0 Safari/537.36",
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "accept-language": "zh-CN,zh;q=0.9",
        }
        for url in urls:
            try:
                async with self._session.get(url, headers=headers,
                                              allow_redirects=True,
                                              timeout=15) as resp:
                    await resp.text()
                    for name, cookie in self._session.cookie_jar.filter_cookies(URL(url)).items():
                        self._cookies[name] = cookie.value
            except Exception:
                pass

        logger.info(f"Cookie收集完成: {len(self._cookies)} 个")

    # ---- 步骤2: 获取 MTOP Token ----
    async def _get_token(self):
        """从H5 API获取 _m_h5_tk token"""
        params = {
            "jsv": "2.7.2",
            "appKey": APP_KEY,
            "t": str(int(time.time() * 1000)),
            "api": "mtop.relationrecommend.WirelessRecommend.recommend",
            "v": "2.0",
            "type": "originaljson",
        }
        try:
            async with self._session.get(BASE_URL, params=params, timeout=15) as resp:
                await resp.text()
                for name, cookie in self._session.cookie_jar.filter_cookies(URL(BASE_URL)).items():
                    self._cookies[name] = cookie.value

            token = self._cookies.get("_m_h5_tk", "")
            if token and "_" in token:
                self._token_part = token.split("_")[0]
                logger.info(f"Token获取成功: {self._token_part[:16]}...")
            else:
                # 尝试从已有cookie中提取
                for k, v in self._cookies.items():
                    if "_" in v and len(v) > 40:
                        self._token_part = v.split("_")[0]
                        logger.info(f"Token从cookie提取: {self._token_part[:16]}...")
                        return
                logger.warning("未获取到有效 _m_h5_tk，使用fallback")
                self._token_part = "fallback"
        except Exception as e:
            logger.warning(f"Token获取失败: {e}，使用fallback")
            self._token_part = "fallback"

    # ---- 步骤3: 图片上传 ----
    async def _upload_image(self, image_path: str) -> Optional[str]:
        """
        上传图片到1688，返回 imageId
        自动转换格式: 任何格式 → JPEG, 缩放到800px
        """
        if not os.path.exists(image_path):
            logger.error(f"图片不存在: {image_path}")
            return None

        try:
            # 加载图片 → 转JPEG → 缩放
            img = Image.open(image_path)
            if img.mode in ("RGBA", "P", "LA"):
                img = img.convert("RGB")
            img.thumbnail((800, 800), Image.LANCZOS)

            buf = BytesIO()
            img.save(buf, format="JPEG", quality=85)
            img_bytes = buf.getvalue()

            import base64
            img_b64 = base64.b64encode(img_bytes).decode()

            logger.debug(f"图片处理: {img.size} → JPEG {len(img_bytes)} bytes")

            # 构造上传请求
            params_data = {
                "searchScene": "imageEx",
                "interfaceName": "imageBase64ToImageId",
                "serviceParam.extendParam[imageBase64]": img_b64,
                "subChannel": "pc_image_search_image_id",
            }
            request_data = {
                "appId": 32517,
                "params": json.dumps(params_data, ensure_ascii=False),
            }
            data_str = json.dumps(request_data, ensure_ascii=False)
            ts = str(int(time.time() * 1000))
            sign = generate_sign(self._token_part, ts, APP_KEY, data_str)

            params = {
                "jsv": "2.7.2", "appKey": APP_KEY, "t": ts, "sign": sign,
                "api": "mtop.relationrecommend.WirelessRecommend.recommend",
                "v": "2.0", "type": "originaljson", "dataType": "jsonp",
                "jsonpIncPrefix": "pf", "timeout": "20000",
            }
            headers = {
                "content-type": "application/x-www-form-urlencoded",
                "origin": "https://s.1688.com",
                "referer": "https://s.1688.com/",
                "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            }

            async with self._session.post(
                BASE_URL, params=params, data={"data": data_str},
                headers=headers, timeout=30
            ) as resp:
                text = await resp.text()
                result = json.loads(text)

                if result.get("data", {}).get("success"):
                    image_id = result["data"]["imageId"]
                    logger.info(f"  图片上传成功 → imageId: {image_id}")
                    return image_id
                else:
                    error_msg = result.get("data", {}).get("errorMsg", "unknown")
                    logger.warning(f"  图片上传失败: {error_msg}")
                    return None

        except Exception as e:
            logger.error(f"  图片上传异常: {e}")
            return None

    # ---- 步骤4: 以图搜款 ----
    async def search_by_image(self, image_path: str) -> List[Product1688]:
        """完整图搜流程: 上传→搜索→解析"""
        # 上传
        image_id = await self._upload_image(image_path)
        if not image_id:
            return []

        # 获取搜索页cookie
        try:
            await self._get_search_page_cookies(image_id, "image")
        except Exception:
            pass

        # 搜索
        params_data = {
            "beginPage": 1, "pageSize": 60,
            "method": "imageOfferSearchService",
            "searchScene": "pcImageSearch",
            "appName": "pctusou", "tab": "imageSearch",
            "imageId": image_id, "imageIdList": image_id,
            "spm": "a26352.13672862.imagesearch.upload",
        }
        request_data = {
            "appId": 32517,
            "params": json.dumps(params_data, ensure_ascii=False),
        }
        data_str = json.dumps(request_data, ensure_ascii=False)
        ts = str(int(time.time() * 1000))
        sign = generate_sign(self._token_part, ts, APP_KEY, data_str)

        params = {
            "jsv": "2.7.2", "appKey": APP_KEY, "t": ts, "sign": sign,
            "api": "mtop.relationrecommend.wirelessrecommend.recommend",
            "v": "2.0", "type": "jsonp", "dataType": "jsonp",
            "timeout": "20000", "jsonpIncPrefix": "pf",
            "callback": f"mtopjsonppf{int(time.time())}",
            "data": data_str,
        }
        cookies_str = "; ".join(f"{k}={v}" for k, v in self._cookies.items())
        headers = {
            "accept": "*/*",
            "accept-language": "zh-CN,zh;q=0.9",
            "cookie": cookies_str,
            "referer": f"https://pages-fast.1688.com/wow/cbu/srch_rec/image_search/youyuan/index.html?tab=imageSearch&imageId={image_id}",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        }

        try:
            async with self._session.get(
                BASE_URL, params=params, headers=headers, timeout=30
            ) as resp:
                raw = await resp.text()
                return self._parse_image_results(raw)
        except Exception as e:
            logger.error(f"  图搜请求异常: {e}")
            return []

    # ---- 步骤5: 关键词搜索 ----
    async def search_by_keyword(self, keyword: str) -> List[Product1688]:
        """关键词搜索1688商品"""
        if not keyword or not keyword.strip():
            return []

        params_data = {
            "beginPage": 1, "pageSize": 30,
            "method": "getOfferList",
            "pageId": "qWJOoeNkRwblv903Iv6KQqPVkYDrgMudKHTRsee9Sjz7N9z1",
            "verticalProductFlag": "pcmarket",
            "searchScene": "pcOfferSearch",
            "charset": "GBK",
            "spm": "a26352.b28411319/2508.searchbox.0",
            "keywords": keyword,
        }
        request_data = {
            "appId": 32517,
            "params": json.dumps(params_data, ensure_ascii=False),
        }
        data_str = json.dumps(request_data, ensure_ascii=False)
        ts = str(int(time.time() * 1000))
        sign = generate_sign(self._token_part, ts, APP_KEY, data_str)

        params = {
            "jsv": "2.7.4", "appKey": APP_KEY, "t": ts, "sign": sign,
            "api": "mtop.relationrecommend.WirelessRecommend.recommend",
            "v": "2.0", "type": "jsonp", "dataType": "jsonp",
            "timeout": "20000", "jsonpIncPrefix": "pf",
            "callback": f"mtopjsonppf{int(time.time())}",
            "data": data_str,
        }
        cookies_str = "; ".join(f"{k}={v}" for k, v in self._cookies.items())
        headers = {
            "accept": "*/*",
            "accept-language": "zh-CN,zh;q=0.9",
            "cookie": cookies_str,
            "referer": f"https://s.1688.com/selloffer/offer_search.htm?keywords={keyword}",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        }

        try:
            async with self._session.get(
                BASE_URL, params=params, headers=headers, timeout=30
            ) as resp:
                raw = await resp.text()
                return self._parse_text_results(raw)
        except Exception as e:
            logger.error(f"  关键词搜索异常: {e}")
            return []

    # ---- 解析: 图搜结果 ----
    def _parse_image_results(self, raw_text: str) -> List[Product1688]:
        """
        解析图搜 JSONP 响应
        商品数据在: data.data.OFFER.items[]
        价格在: trackInfo.expoData (^price@xx.xx)
        """
        products = []
        try:
            # 剥离 JSONP wrapper
            start = raw_text.find("(")
            end = raw_text.rfind(")")
            if start > 0 and end > 0:
                json_str = raw_text[start + 1:end]
            else:
                json_str = raw_text

            result = json.loads(json_str)
            items = (
                result.get("data", {}).get("data", {}).get("OFFER", {}).get("items", [])
            )
            if not items:
                # 尝试其他路径
                items = result.get("data", {}).get("OFFER", {}).get("items", [])

            logger.debug(f"  图搜解析: {len(items)} 条item")

            for item in items:
                p = self._parse_image_item(item)
                if p and p.price_low > 0:
                    products.append(p)

        except json.JSONDecodeError as e:
            logger.warning(f"  图搜JSON解析失败: {e}")
        except Exception as e:
            logger.warning(f"  图搜解析异常: {e}")

        # 去重 + 按价格排序
        return self._dedup_and_sort(products)

    def _parse_image_item(self, item: dict) -> Optional[Product1688]:
        """从单个图搜结果item提取商品信息"""
        data = item.get("data", {})
        expo = item.get("trackInfo", {}).get("expoData", "")

        # 从 expoData 提取价格: ^price@10.50
        price_match = re.search(r'\^price@([\d.]+)', expo)
        price = float(price_match.group(1)) if price_match else 0.0

        # 从 expoData 提取 offerId: ^object_id@12345678
        oid_match = re.search(r'\^object_id@(\d+)', expo)
        offer_id = oid_match.group(1) if oid_match else ""

        # 从 expoData 提取标题: professionPV@ 或 yoloCropRegion@ 后面的商品描述
        title = ""
        for field in ("professionPV@", "filterInfo@"):
            pv_match = re.search(rf'{field}([^^]+)', expo)
            if pv_match:
                raw = pv_match.group(1)
                # 清理: 只保留中文、字母、数字、常用符号
                cleaned = re.sub(r'[^一-鿿\w\s:：;；,，.。\-()（）]', '', raw)
                cleaned = cleaned.strip(';:：,， ')
                if cleaned and len(cleaned) > 3:
                    title = cleaned
                    break

        # loginId 是店铺名
        shop_name_raw = data.get("loginId", "")
        # 尝试GBK解码店铺名
        shop_name = shop_name_raw
        try:
            shop_name = shop_name_raw.encode('latin-1').decode('gbk')
        except (UnicodeDecodeError, UnicodeEncodeError):
            pass
        province = data.get("province", "")
        pic_url = data.get("offerPicUrl", "")

        # 如果没解析到标题，用店铺名+所在地
        display_title = title if title else shop_name

        if not offer_id and not display_title:
            return None

        return Product1688(
            offer_id=offer_id,
            title=display_title,
            price_low=price,
            price_high=price,
            shop_name=shop_name,
            location=province,
            detail_url=f"https://detail.1688.com/offer/{offer_id}.html" if offer_id else "",
            image_url=pic_url,
        )

    # ---- 解析: 文字搜索结果 ----
    def _parse_text_results(self, raw_text: str) -> List[Product1688]:
        """解析文字搜索JSONP响应"""
        products = []
        try:
            start = raw_text.find("(")
            end = raw_text.rfind(")")
            json_str = raw_text[start + 1:end] if start > 0 and end > 0 else raw_text

            result = json.loads(json_str)
            items = (
                result.get("data", {}).get("data", {}).get("OFFER", {}).get("items", [])
            )

            for item in items:
                data = item.get("data", {})
                price_info = data.get("priceInfo", {})

                price_str = str(price_info.get("price", "0")) if isinstance(price_info, dict) else "0"
                try:
                    price = float(price_str)
                except (ValueError, TypeError):
                    price = 0.0

                shop = data.get("shopAddition", {})
                shop_name = shop.get("text", "") if isinstance(shop, dict) else ""

                offer_id = str(data.get("offerId", data.get("id", "")))
                title = data.get("title", "")
                pic = data.get("offerPicUrl", data.get("imgUrl", ""))
                province = data.get("province", "")
                sales = data.get("saleQuantity", data.get("bookedCount", 0))

                if price > 0 or title:
                    products.append(Product1688(
                        offer_id=offer_id,
                        title=title,
                        price_low=price,
                        price_high=price,
                        shop_name=shop_name,
                        location=province,
                        detail_url=f"https://detail.1688.com/offer/{offer_id}.html" if offer_id else "",
                        image_url=pic,
                        sales=int(sales) if sales else 0,
                    ))
        except Exception as e:
            logger.warning(f"  文字搜索解析异常: {e}")

        return self._dedup_and_sort(products)

    # ---- 工具 ----
    @staticmethod
    def _dedup_and_sort(products: List[Product1688]) -> List[Product1688]:
        seen = set()
        unique = []
        for p in products:
            key = p.offer_id or p.title[:40]
            if key and key not in seen:
                seen.add(key)
                unique.append(p)
        unique.sort(key=lambda x: x.price_low if x.price_low > 0 else 999999)
        return unique[:20]

    async def _get_search_page_cookies(self, search_param: str, search_type: str):
        """访问搜索页获取额外cookie"""
        if search_type == "image":
            url = f"https://s.1688.com/youyuan/index.htm?tab=imageSearch&imageId={search_param}&imageIdList={search_param}"
        else:
            url = f"https://s.1688.com/selloffer/offer_search.htm?keywords={search_param}"

        headers = {
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }
        try:
            async with self._session.get(url, headers=headers,
                                          allow_redirects=True, timeout=15) as resp:
                await resp.text()
                for name, cookie in self._session.cookie_jar.filter_cookies(URL(url)).items():
                    self._cookies[name] = cookie.value
        except Exception:
            pass


# ============================================================
# 数据加载
# ============================================================
def load_translations() -> Dict[str, str]:
    """加载俄语→中文翻译CSV"""
    translations = {}
    if CSV_TRANSLATION.exists():
        with open(CSV_TRANSLATION, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if len(row) >= 2 and row[0].strip() and row[1].strip():
                    translations[row[0].strip()] = row[1].strip()
    logger.info(f"翻译词库: {len(translations)} 条")
    return translations


def lookup_translation(text: str, trans: Dict[str, str]) -> str:
    """查找翻译，精确+前缀模糊匹配"""
    if not text:
        return ""
    if text in trans:
        return trans[text]
    for key, val in trans.items():
        if key.startswith(text) or text.startswith(key):
            return val
    return ""


async def download_image(session: aiohttp.ClientSession, url: str,
                         local_path: Path) -> bool:
    """下载单张图片"""
    if local_path.exists() and local_path.stat().st_size > 100:
        return True
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://www.ozon.ru/",
        }
        async with session.get(url, headers=headers, timeout=20) as resp:
            if resp.status == 200:
                local_path.write_bytes(await resp.read())
                return True
    except Exception:
        pass
    return False


async def load_products(excel_path: str, limit: int = 0,
                        start_index: int = 0) -> List[dict]:
    """加载Excel商品数据 + 翻译 + 下载图片"""
    translations = load_translations()

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    products = []

    for r in range(2, ws.max_row + 1):
        title_ru = str(ws.cell(row=r, column=4).value or "").strip()
        if not title_ru:
            continue

        image_url = str(ws.cell(row=r, column=2).value or "").strip()

        products.append({
            "row": r - 1,
            "image_url": image_url,
            "title_ru": title_ru,
            "title_cn": lookup_translation(title_ru, translations),
            "brand": str(ws.cell(row=r, column=7).value or "").strip(),
            "category": str(ws.cell(row=r, column=8).value or "").strip(),
            "price_rub": str(ws.cell(row=r, column=10).value or "").strip(),
            "sales": str(ws.cell(row=r, column=11).value or "").strip(),
        })

        if limit and len(products) >= limit:
            break

    wb.close()

    # 下载图片
    logger.info(f"读取 {len(products)} 个商品，开始下载图片...")
    downloaded = 0
    async with aiohttp.ClientSession() as session:
        for p in products:
            if not p["image_url"]:
                continue
            name_hash = hashlib.md5(p["image_url"].encode()).hexdigest()[:12]
            local = IMAGE_DIR / f"{name_hash}.jpg"
            ok = await download_image(session, p["image_url"], local)
            p["image_local"] = str(local) if (ok or local.exists()) else ""
            if ok:
                downloaded += 1

    logger.info(f"图片就绪: {downloaded}/{len(products)}")
    return products


# ============================================================
# 利润计算
# ============================================================
def calculate_profit(ozon_price_rub: str, cost_price_cny: float,
                     weight_kg: float = 0.3) -> dict:
    """
    计算利润

    收入 = Ozon售价(RUB) × 汇率
    成本 = 1688批发价 + 国内运费 + 国际物流 + Ozon佣金
    国际物流简化估算: 30元/kg (小包)
    """
    # 清理价格字符串: 去掉 ₽、空格、逗号等
    clean_price = str(ozon_price_rub).replace("₽", "").replace(" ", "").replace(",", ".").strip()
    try:
        price_rub = float(clean_price)
    except (ValueError, TypeError):
        price_rub = 0.0

    revenue_cny = price_rub * RUB_TO_CNY
    intl_shipping = weight_kg * 30.0          # 国际物流估算
    commission = revenue_cny * OZON_COMMISSION_RATE
    total_cost = cost_price_cny + DOMESTIC_SHIPPING + intl_shipping + commission
    net_profit = revenue_cny - total_cost

    if revenue_cny > 0:
        margin = (net_profit / revenue_cny) * 100
    else:
        margin = 0.0

    return {
        "revenue_cny": round(revenue_cny, 2),
        "cost_1688": round(cost_price_cny, 2),
        "cost_domestic": DOMESTIC_SHIPPING,
        "cost_intl_shipping": round(intl_shipping, 2),
        "cost_commission": round(commission, 2),
        "total_cost": round(total_cost, 2),
        "net_profit": round(net_profit, 2),
        "margin_pct": round(margin, 1),
    }


# ============================================================
# 断点保存
# ============================================================
def load_checkpoint() -> set:
    """加载已完成的row号"""
    if CHECKPOINT_FILE.exists():
        with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return {r["row"] for r in data if r.get("status") == "success"}
    return set()


def save_checkpoint(results: List[SearchResult]):
    """每次搜索完保存"""
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump([asdict(r) for r in results], f, ensure_ascii=False, indent=2)


# ============================================================
# 主流程
# ============================================================
async def run_pipeline(products: List[dict], use_image: bool = True,
                       use_keyword: bool = True, start_index: int = 0):
    """执行搜索管道"""
    results: List[SearchResult] = []
    completed = load_checkpoint()
    if completed:
        logger.info(f"断点恢复: 已跳过 {len(completed)} 个已完成商品")

    async with H5Searcher() as searcher:
        for i, p in enumerate(products):
            if i < start_index:
                continue

            row = p["row"]
            if row in completed:
                logger.info(f"[{i+1}/{len(products)}] 第{row}行 ⏭️ 已跳过")
                continue

            title_ru = p["title_ru"]
            cn = p["title_cn"]
            image_local = p.get("image_local", "")

            logger.info(f"\n{'='*60}")
            logger.info(f"[{i+1}/{len(products)}] 第{row}行")
            logger.info(f"  俄语: {title_ru[:60]}")
            logger.info(f"  中文: {cn[:40] if cn else '(无翻译)'}")

            result = SearchResult(
                row=row,
                ozon_title=title_ru,
                ozon_title_cn=cn,
                ozon_price_rub=p["price_rub"],
                ozon_sales=p["sales"],
                ozon_image_url=p["image_url"],
                category=p["category"],
                brand=p["brand"],
                image_local=image_local,
            )

            all_matches: List[Product1688] = []

            # 图片搜索
            if use_image and image_local and os.path.exists(image_local):
                logger.info("  📷 以图搜款...")
                try:
                    matches = await searcher.search_by_image(image_local)
                    all_matches.extend(matches)
                    logger.info(f"  图搜结果: {len(matches)} 个")
                except Exception as e:
                    logger.warning(f"  图搜异常: {e}")

            # 关键词搜索兜底
            if use_keyword and cn:
                await asyncio.sleep(1.0)  # 间隔
                logger.info(f"  🔍 关键词搜索: {cn[:30]}")
                try:
                    matches = await searcher.search_by_keyword(cn)
                    all_matches.extend(matches)
                    logger.info(f"  关键词结果: {len(matches)} 个")
                except Exception as e:
                    logger.warning(f"  关键词搜索异常: {e}")

            # 合并去重排序
            if all_matches:
                seen = set()
                unique = []
                for m in all_matches:
                    key = m.offer_id or m.title[:30]
                    if key and key not in seen:
                        seen.add(key)
                        unique.append(m)
                unique.sort(key=lambda x: x.price_low if x.price_low > 0 else 999999)
                result.products = [asdict(m) for m in unique[:10]]
                result.status = "success"
                logger.info(f"  ✅ 共匹配 {len(unique)} 个商品:")
                for m in unique[:5]:
                    logger.info(f"     ¥{m.display_price} | {m.title[:35]} | {m.shop_name[:15]}")
            else:
                result.status = "no_results"
                logger.info("  ⚠️ 无匹配结果")

            results.append(result)
            save_checkpoint(results)

            # 商品间间隔
            await asyncio.sleep(1.5)

    return results


# ============================================================
# Excel 输出
# ============================================================
def save_excel(results: List[SearchResult], output_path: str):
    """格式化输出Excel + 利润计算"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1688利润核算"

    headers = [
        "序号", "Ozon标题(俄语)", "中文关键词", "Ozon售价(RUB)",
        "Ozon销量", "品牌", "类目",
        "1688商品标题", "1688批发价(¥)", "起批量", "店铺",
        "所在地", "1688链接",
        "收入(¥)", "1688成本(¥)", "国内运费(¥)", "国际物流(¥)",
        "Ozon佣金(¥)", "总成本(¥)", "净利润(¥)", "毛利率",
        "状态", "备注",
    ]
    widths = [5, 35, 18, 9, 7, 9, 15, 35, 11, 7, 14, 8, 30,
              9, 10, 8, 8, 9, 9, 9, 7, 8, 16]

    # 样式
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="微软雅黑", size=9)
    link_font = Font(name="微软雅黑", size=9, color="0563C1", underline="single")
    profit_font = Font(name="微软雅黑", size=9, bold=True, color="006100")
    loss_font = Font(name="微软雅黑", size=9, bold=True, color="C00000")
    border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # 写表头
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    row_idx = 2
    for r in results:
        base = [
            r.row, r.ozon_title, r.ozon_title_cn,
            r.ozon_price_rub.replace("₽", "").strip(),
            r.ozon_sales, r.brand, r.category,
        ]

        if r.products:
            for p in r.products:
                cost_price = p.get("price_low", 0) or 0
                profit = calculate_profit(r.ozon_price_rub, cost_price)

                vals = base + [
                    p.get("title", ""),
                    p.get("price_text", "") or f"{cost_price:.2f}",
                    p.get("min_order", ""),
                    p.get("shop_name", ""),
                    p.get("location", ""),
                    p.get("detail_url", ""),
                    profit["revenue_cny"],
                    profit["cost_1688"],
                    profit["cost_domestic"],
                    profit["cost_intl_shipping"],
                    profit["cost_commission"],
                    profit["total_cost"],
                    profit["net_profit"],
                    f"{profit['margin_pct']}%",
                    r.status,
                    r.error_msg,
                ]
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=ci, value=v)
                    cell.font = cell_font
                    cell.border = border
                    cell.fill = green_fill
                    cell.alignment = Alignment(vertical="center")
                    # 链接
                    if ci == 13 and isinstance(v, str) and v.startswith("http"):
                        cell.font = link_font
                        cell.hyperlink = v
                    # 净利润颜色
                    if ci == 21:
                        try:
                            val = float(v) if v else 0
                            cell.font = profit_font if val > 0 else loss_font
                        except (ValueError, TypeError):
                            pass
                row_idx += 1
        else:
            vals = base + ["", "", "", "", "", ""] + [""] * 8 + [r.status, r.error_msg]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.font = cell_font
                cell.border = border
                cell.fill = red_fill
            row_idx += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"

    wb.save(output_path)
    wb.close()
    logger.info(f"📊 结果已保存: {output_path} ({row_idx - 2} 行)")


# ============================================================
# CLI
# ============================================================
async def main():
    parser = argparse.ArgumentParser(
        description="1688 以图搜商 + 利润核算",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python profit_finder.py --limit 3               # 测试3条(图搜+关键词)
  python profit_finder.py --limit 10 --no-image   # 仅关键词搜索
  python profit_finder.py --no-keyword            # 仅图片搜索
  python profit_finder.py --start 5 --limit 10    # 断点续跑
        """,
    )
    parser.add_argument("--limit", type=int, default=0, help="限制商品数(0=全部)")
    parser.add_argument("--start", type=int, default=0, help="起始索引")
    parser.add_argument("--no-image", action="store_true", help="跳过图片搜索")
    parser.add_argument("--no-keyword", action="store_true", help="跳过关键词搜索")
    parser.add_argument("-i", "--input", default="", help="输入Excel路径")
    parser.add_argument("-o", "--output", default="", help="输出Excel路径")
    args = parser.parse_args()

    excel_path = args.input or str(DEFAULT_INPUT)
    output_path = args.output or str(DEFAULT_OUTPUT)

    if not os.path.exists(excel_path):
        logger.error(f"输入文件不存在: {excel_path}")
        return

    # 加载数据
    products = await load_products(excel_path, limit=args.limit, start_index=args.start)
    if not products:
        logger.error("没有商品数据")
        return

    # 搜索
    results = await run_pipeline(
        products=products,
        use_image=not args.no_image,
        use_keyword=not args.no_keyword,
        start_index=args.start,
    )

    # 输出
    if results:
        save_excel(results, output_path)
        stats = {
            "total": len(results),
            "success": sum(1 for r in results if r.status == "success"),
            "no_results": sum(1 for r in results if r.status == "no_results"),
            "error": sum(1 for r in results if r.status == "error"),
        }
        logger.info(f"\n{'='*60}")
        logger.info(f"✅ 完成! 成功: {stats['success']}/{stats['total']} | "
                    f"无结果: {stats['no_results']} | 异常: {stats['error']}")
        logger.info(f"📊 输出: {output_path}")

        # 利润摘要
        total_profit = 0.0
        profitable_count = 0
        for r in results:
            for p in r.products:
                cost = p.get("price_low", 0) or 0
                profit = calculate_profit(r.ozon_price_rub, cost)
                total_profit += profit["net_profit"]
                if profit["net_profit"] > 0:
                    profitable_count += 1
                break  # 每个商品只算第一个匹配

        logger.info(f"💰 预估总利润: ¥{total_profit:.2f} | "
                    f"盈利商品: {profitable_count}/{stats['success']}")


if __name__ == "__main__":
    asyncio.run(main())
